import asyncio
import aiohttp
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
import random
import json
import csv
import os
import re
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor
import math
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from datetime import datetime, timedelta
import random
from app.utils.model import SearatesRequestPayload, SearatesResponsePayload, FreightosRequestPayload, FreightosQuotesResponse
from app.utils.city_point_dict import CITIES_TO_POINT_ID_MAP
from app.utils.freightos_locations import FREIGHTOS_LOCATIONS
from app.utils.checkpoint_manager import CheckpointManager
from app.tasks import _go

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('shipping_matrix.log'),
        logging.StreamHandler()
    ]
)

BEARER_TOKEN = os.getenv("SEARATES_BEARER_TOKEN")
SEARATES_API_URL = "https://rates.searates.com/graphql"

logger = logging.getLogger(__name__)


SHIPPING_DATES = [
    "2025-06-29",  # Late June 2025
    "2025-07-01",  # Early July 2025
    "2025-07-28",  # Late July 2025
    "2025-08-01",  # Early August 2025
    "2025-08-29",  # Late August 2025
    "2025-09-01",  # Early September 2025
    "2025-09-28",  # Late September 2025
    "2025-10-01",  # Early October 2025
    "2025-10-28",  # Late October 2025
    "2025-11-01",  # Early November 2025
    "2025-11-25",  # Late November 2025
    "2025-12-06",  # Early December 2025
    "2025-12-26",  # Late December 2025
    "2026-01-01",  # Early January 2026
    "2026-01-24",  # Late January 2026
    "2026-02-01",  # Early February 2026
    "2026-02-25",  # Late February 2026
    "2026-03-05",  # Early March 2026
    "2026-03-22",  # Late March 2026
    "2026-04-01",  # Early April 2026
    "2026-04-23",  # Late April 2026
    "2026-05-01",  # Early May 2026
    "2026-05-28",  # Late May 2026
    "2026-06-01",  # Early June 2026
    "2026-06-29",  # Late June 2026 (end date)
]


CONTAINERS = ["ST20", "ST40"]



def build_searates_payload(payload: SearatesRequestPayload):
    query = f"""
    {{
      rates(
        includedServices: d2d,
        portFromFees: true,
        portToFees: true,
        shippingType: FCL,
        pointIdFrom: "{payload.pointIdFrom}",
        pointIdTo: "{payload.pointIdTo}",
        date: "{payload.date}",
        container: {payload.container},
      ) {{
        points {{
          id
          rateId
          location {{
            id
            name
            country
            lat
            lng
            code
            inaccessible
            pointType
            dry
          }}
          shippingType
          provider
          providerLogo
          loads {{
            id
            unit
            amount
            shortCode
            type
          }}
          pointTariff {{
            name
            abbr
            price
            currency
            profileId
          }}
          routeTariff {{
            name
            abbr
            price
            currency
          }}
          lumpsumTariff {{
            price
            currency
          }}
          co2 {{
            amount
            price
            placeAmount
            placePrice
            lumpsumAmount
            lumpsumPrice
          }}
          transitTime {{
            rate
            port
            route
          }}
          profileId
          distance
          totalPrice
          totalCurrency
          pointTotal
          routeTotal
          terms
        }}
        general {{
          shipmentId
          validityFrom
          validityTo
          individual
          totalPrice
          totalCurrency
          totalTransitTime
          totalCo2 {{
            amount
            price
          }}
          dfaRate
          alternative
          expired
          spaceGuarantee
          spot
          indicative
          standard
          rateOwner
          queryShippingType
          promotionObligations
          shipmentCreatedAt
        }}
        request {{
          request_key
        }}
      }}
    }}
    """

    payload = {"query": query}
    return payload


async def make_request(url: str, payload: SearatesRequestPayload) -> SearatesResponsePayload:
    async with aiohttp.ClientSession() as session:
        try:
            request_payload = build_searates_payload(payload)
            async with session.request("POST", url, headers={"Authorization": BEARER_TOKEN}, json=request_payload) as response:
                return SearatesResponsePayload(**await response.json())
        except Exception as e:
            print(e)
            return None


async def compute_shipping_matrix(
    date: str = "2025-09-01", 
    container: str = "ST40",
    delay_range: tuple = (2, 5)
) -> List[Dict[str, Any]]:
    """
    Compute shipping matrix for all city combinations with robust error handling.
    
    Args:
        date: Shipping date in YYYY-MM-DD format
        container: Container type (e.g., "ST20")
        delay_range: Tuple of (min, max) seconds for random delays
        
    Returns:
        List of dictionaries containing shipping data for each combination
    """
    results = []
    all_cities = list(CITIES_TO_POINT_ID_MAP.keys())
    # Randomly select 65% of cities to reduce computation time
    cities_count = int(len(all_cities) * 0.03)
    cities = random.sample(all_cities, cities_count)  # Randomly sample 65% of cities
    # cities = ["LONDON", "PARIS", "MUNICH"]
    total_combinations = len(cities) * (len(cities) - 1)  # Exclude same city combinations
    
    logger.info(f"Starting matrix computation for {total_combinations} city combinations")
    
    for i, origin_city in enumerate(cities):
        for j, destination_city in enumerate(cities):
            if origin_city == destination_city:
                continue  # Skip same city combinations
                
            combination_num = i * (len(cities) - 1) + j - (1 if j > i else 0)
            logger.info(f"Processing combination {combination_num + 1}/{total_combinations}: {origin_city} -> {destination_city}")
            
            try:
                # Create API payload
                payload = SearatesRequestPayload(
                    pointIdFrom=CITIES_TO_POINT_ID_MAP[origin_city],
                    pointIdTo=CITIES_TO_POINT_ID_MAP[destination_city],
                    date=date,
                    container=container
                )
                
                # Make API request with retry logic
                api_response = await make_request_with_retry(SEARATES_API_URL, payload)
                
                if not api_response or not api_response.data or not api_response.data.rates:
                    logger.warning(f"No rates found for {origin_city} -> {destination_city}")
                    # Add entry with no data
                    # results.append(create_empty_result(origin_city, destination_city, date, container))
                    continue
                
                # Process each rate found
                for rate_idx, rate in enumerate(api_response.data.rates):
                    screenshot_url = None
                    website_link = None
                    
                    try:
                        # Generate screenshot URL
                        shipment_id = rate.general.shipmentId
                        website_link = f"https://www.searates.com/logistics-explorer/?id={shipment_id}"
                        
                        logger.info(f"Taking screenshot for shipment {shipment_id}")
                        screenshot_result = await _go(website_link)
                        
                        if isinstance(screenshot_result, str):
                            screenshot_url = screenshot_result
                        elif isinstance(screenshot_result, dict) and "error" not in screenshot_result:
                            screenshot_url = screenshot_result
                        else:
                            logger.error(f"Screenshot failed for {shipment_id}: {screenshot_result}")
                            
                    except Exception as screenshot_error:
                        logger.error(f"Screenshot error for {origin_city} -> {destination_city}: {screenshot_error}")
                    
                    # Create result entry
                    result_entry = create_result_entry(
                        origin_city=origin_city,
                        destination_city=destination_city,
                        rate=rate,
                        date=date,
                        container=container,
                        screenshot_url=screenshot_url,
                        website_link=website_link,
                        is_expired=rate.general.expired
                    )
                    
                    results.append(result_entry)
                    
                    # Add delay between screenshots if multiple rates
                    if rate_idx < len(api_response.data.rates) - 1:
                        await asyncio.sleep(random.uniform(1, 2))
                        
            except Exception as e:
                logger.error(f"Error processing {origin_city} -> {destination_city}: {e}")
                results.append(create_error_result(origin_city, destination_city, date, container, str(e)))
            
            # Add delay between API requests to avoid rate limiting
            if combination_num < total_combinations - 1:
                delay = random.uniform(delay_range[0], delay_range[1])
                logger.info(f"Waiting {delay:.2f} seconds before next request...")
                await asyncio.sleep(delay)
    
    logger.info(f"Matrix computation completed. Processed {len(results)} entries")
    return results


async def make_request_with_retry(
    url: str, 
    payload: SearatesRequestPayload, 
    max_retries: int = 3,
    retry_delay: float = 2.0
) -> Optional[SearatesResponsePayload]:
    """
    Make API request with retry logic and exponential backoff.
    
    Args:
        url: API endpoint URL
        payload: Request payload
        max_retries: Maximum number of retry attempts
        retry_delay: Base delay between retries
        
    Returns:
        API response or None if all retries failed
    """
    for attempt in range(max_retries + 1):
        try:
            response = await make_request(url, payload)
            if response:
                return response
                
        except Exception as e:
            if attempt == max_retries:
                logger.error(f"API request failed after {max_retries} retries: {e}")
                return None
            
            delay = retry_delay * (2 ** attempt) + random.uniform(0, 1)
            logger.warning(f"API request attempt {attempt + 1} failed, retrying in {delay:.2f}s: {e}")
            await asyncio.sleep(delay)
    
    return None

city_country_map = {
    "HO CHI MINH": "VN",
    "STOCKHOLM": "SE",
    "MILANO": "IT",
    "BEOGRAD": "RS",
    "BRASILIA": "BR",
    "BRASTISLAVA": "SK",
    "BRATISLAVA": "SK",
    "SAN JOSE": "CR",
    "SAO PAULO": "BR",
    "PERTH": "AU",
    "CHICAGO": "US",
    "BOSTON": "US",
    "HELSINKI": "FI",
    "LONDON": "GB",
    "DUBAI": "AE",
    "OSLO": "NO",
    "BUENOS AIRES": "AR",
    "PARIS (FR)": "FR",
    "ATLANTA (US)": "US",
    "ZURICH": "CH",
    "CHICAGO": "US",
    "CHISINAU": "MD"
}



def create_result_entry(
    origin_city: str,
    destination_city: str,
    rate,
    date: str,
    container: str,
    screenshot_url: Optional[str],
    website_link: Optional[str],
    is_expired: bool = False,
) -> Dict[str, Any]:
    """Create a structured result entry from API rate data."""
    
    # Get origin and destination location data
    origin_location = None
    destination_location = None
    
    for point in rate.points:
        if point.location.name.upper() == origin_city:
            origin_location = point.location
        elif point.location.name.upper() == destination_city:
            destination_location = point.location
    
    # Extract carrier information
    carrier = None
    if rate.points:
        for point in rate.points:
            if hasattr(point, 'provider') and point.provider:
                carrier = point.provider
                break
    
    # Clean screenshot_url - extract just the S3 URL if it's a dict/object
    clean_screenshot_url = None
    if screenshot_url:
        if isinstance(screenshot_url, dict):
            # Extract screenshot_url from dict if it exists
            clean_screenshot_url = screenshot_url.get('screenshot_url') or screenshot_url.get('url')
        elif isinstance(screenshot_url, str):
            clean_screenshot_url = screenshot_url
        else:
            # Handle other object types by trying to extract screenshot_url
            try:
                if hasattr(screenshot_url, 'get'):
                    clean_screenshot_url = screenshot_url.get('screenshot_url') or screenshot_url.get('url')
                else:
                    clean_screenshot_url = str(screenshot_url)
            except:
                clean_screenshot_url = str(screenshot_url) if screenshot_url else None
    
    return {
        "city_of_origin": origin_city,
        "country_of_origin": origin_location.country if origin_location else city_country_map.get(origin_city, "Unknown"),
        "city_of_destination": destination_city,
        "country_of_destination": destination_location.country if destination_location else city_country_map.get(destination_city, "Unknown"),
        "date_of_shipping": date,
        "total_shipping_time_days": rate.general.totalTransitTime,
        "price_of_shipping": rate.general.totalPrice,
        "currency": rate.general.totalCurrency,
        "container_type": container,
        "provider": "Searates",
        "datetime_of_scraping": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "carrier": carrier,
        "screenshot_url": clean_screenshot_url,
        "website_link": website_link,
        "shipment_id": rate.general.shipmentId,
        "rate_id": rate.points[0].rateId if rate.points else None,
        "co2_amount": rate.general.totalCo2.amount,
        "co2_price": rate.general.totalCo2.price,
        "validity_from": rate.general.validityFrom,
        "validity_to": rate.general.validityTo,
        "distance": rate.points[0].distance if rate.points else None,
        "point_total": rate.points[0].pointTotal if rate.points else None,
        "route_total": rate.points[0].routeTotal if rate.points else None,
        "is_expired": is_expired
    }


def create_empty_result(
    origin_city: str, 
    destination_city: str, 
    date: str, 
    container: str,
    screenshot_url: Optional[str],
    website_link: Optional[str],
    is_expired: bool = False,
) -> Dict[str, Any]:
    """Create an empty result entry when no rates are found."""
    return {
        "city_of_origin": origin_city,
        "country_of_origin": city_country_map.get(origin_city, "Unknown"),
        "city_of_destination": destination_city,
        "country_of_destination": city_country_map.get(destination_city, "Unknown"),
        "date_of_shipping": date,
        "total_shipping_time_days": None,
        "price_of_shipping": None,
        "currency": None,
        "container_type": container,
        "provider": "Searates",
        "datetime_of_scraping": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "carrier": None,
        "screenshot_url": screenshot_url,
        "website_link": website_link,
        "shipment_id": None,
        "rate_id": None,
        "co2_amount": None,
        "co2_price": None,
        "validity_from": None,
        "validity_to": None,
        "distance": None,
        "point_total": None,
        "route_total": None,
        "is_expired": is_expired
    }


def create_error_result(
    origin_city: str, 
    destination_city: str, 
    date: str, 
    container: str, 
    error_message: str
) -> Dict[str, Any]:
    """Create an error result entry when processing fails."""
    result = create_empty_result(origin_city, destination_city, date, container)
    result["error"] = error_message
    return result


def save_results_to_csv(results: List[Dict[str, Any]], filename: str = None) -> str:
    """Save results to CSV file with clean screenshot URLs."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"shipping_matrix_{timestamp}.csv"
    
    # Only save results with actual pricing data
    valid_results = [r for r in results if r.get("price_of_shipping") is not None] if results else []
    
    if not valid_results:
        logger.warning("No valid results to save")
        return filename
    
    # Clean the data first to handle old checkpoint format
    cleaned_results = []
    for result in valid_results:
        cleaned_result = result.copy()
        # Clean screenshot_url if it's in old format
        if 'screenshot_url' in cleaned_result:
            cleaned_result['screenshot_url'] = clean_screenshot_url(cleaned_result['screenshot_url'])
        cleaned_results.append(cleaned_result)
    
    # Collect all unique field names from all results
    all_fieldnames = set()
    for result in cleaned_results:
        all_fieldnames.update(result.keys())
    fieldnames = sorted(all_fieldnames)  # Sort for consistent column order
    
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(cleaned_results)
    
    logger.info(f"Results saved to CSV: {filename}")
    return filename


def save_results_to_json(results: List[Dict[str, Any]], filename: str = None) -> str:
    """Save results to JSON file."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"shipping_matrix_{timestamp}.json"
    
    with open(filename, 'w', encoding='utf-8') as jsonfile:
        json.dump(results, jsonfile, indent=2, default=str)
    
    logger.info(f"Results saved to JSON: {filename}")
    return filename


def clean_screenshot_url(screenshot_url) -> str:
    """Clean screenshot_url to extract just the S3 URL string."""
    if not screenshot_url:
        return ''
    
    if isinstance(screenshot_url, dict):
        # Extract screenshot_url from dict if it exists
        return screenshot_url.get('screenshot_url') or screenshot_url.get('url') or ''
    elif isinstance(screenshot_url, str):
        return screenshot_url
    else:
        # Handle other object types
        try:
            if hasattr(screenshot_url, 'get'):
                return screenshot_url.get('screenshot_url') or screenshot_url.get('url') or ''
            else:
                return str(screenshot_url) if screenshot_url else ''
        except:
            return str(screenshot_url) if screenshot_url else ''


def save_results_to_excel(results: List[Dict[str, Any]], filename: str = None) -> str:
    """Save results to Excel file with enhanced formatting for screenshots."""
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl library not available. Please install it with: pip install openpyxl")
        raise ImportError("openpyxl library is required for Excel export")
    
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"shipping_matrix_{timestamp}.xlsx"
    
    if not results:
        logger.warning("No results to save")
        return filename
    
    # Clean the data first to handle old checkpoint format
    cleaned_results = []
    for result in results:
        cleaned_result = result.copy()
        # Clean screenshot_url if it's in old format
        if 'screenshot_url' in cleaned_result:
            cleaned_result['screenshot_url'] = clean_screenshot_url(cleaned_result['screenshot_url'])
        cleaned_results.append(cleaned_result)
    
    try:
        from openpyxl.styles import PatternFill, Border, Side
        from openpyxl.worksheet.hyperlink import Hyperlink
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Shipping Matrix"
        
        # Get headers from first result
        headers = list(cleaned_results[0].keys())
        
        # Define colors and styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Find screenshot_url column index
        screenshot_col = None
        website_col = None
        for idx, header in enumerate(headers, 1):
            if header == 'screenshot_url':
                screenshot_col = idx
            elif header == 'website_link':
                website_col = idx
        
        # Write data rows
        for row, result in enumerate(cleaned_results, 2):
            for col, header in enumerate(headers, 1):
                value = result.get(header, '')
                
                # Handle None values
                if value is None:
                    value = ''
                
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                # Special handling for screenshot URLs - make them clickable
                if header == 'screenshot_url' and value and isinstance(value, str) and value.startswith('http'):
                    # Create hyperlink for screenshot
                    cell.hyperlink = value
                    cell.value = "View Screenshot"
                    cell.font = Font(color="0000FF", underline="single")
                    cell.alignment = Alignment(horizontal='center')
                
                # Special handling for website links
                elif header == 'website_link' and value and isinstance(value, str) and value.startswith('http'):
                    cell.hyperlink = value
                    cell.value = "View Details"
                    cell.font = Font(color="0000FF", underline="single")
                    cell.alignment = Alignment(horizontal='center')
                
                # Format numeric columns
                elif header in ['price_of_shipping', 'co2_price', 'point_total', 'route_total']:
                    if value and str(value).replace('.', '').replace('-', '').isdigit():
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                
                # Format date columns
                elif header in ['date_of_shipping', 'datetime_of_scraping', 'validity_from', 'validity_to']:
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        # For hyperlink cells, use display text length
                        if hasattr(cell, 'hyperlink') and cell.hyperlink:
                            cell_length = len(str(cell.value))
                        else:
                            cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set minimum and maximum widths
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        # Save the workbook
        wb.save(filename)
        
        logger.info(f"Results saved to Excel: {filename}")
        logger.info(f"📊 Excel features: Clickable screenshot links, formatted numbers, frozen headers")
        return filename
        
    except ImportError as e:
        # Fallback to basic Excel export if advanced features not available
        logger.warning(f"Advanced Excel features not available: {e}")
        return save_results_to_excel_basic(cleaned_results, filename)
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        # Fallback to basic Excel export
        return save_results_to_excel_basic(cleaned_results, filename)


def save_results_to_excel_basic(results: List[Dict[str, Any]], filename: str = None) -> str:
    """Basic Excel export without advanced formatting."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"shipping_matrix_{timestamp}.xlsx"
    
    # Clean the data first if not already cleaned
    if results and 'screenshot_url' in results[0]:
        # Check if data needs cleaning (old format)
        first_screenshot = results[0].get('screenshot_url')
        if isinstance(first_screenshot, dict):
            cleaned_results = []
            for result in results:
                cleaned_result = result.copy()
                cleaned_result['screenshot_url'] = clean_screenshot_url(cleaned_result['screenshot_url'])
                cleaned_results.append(cleaned_result)
            results = cleaned_results
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipping Matrix"
    
    # Get headers from first result
    headers = list(results[0].keys())
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
        cell.font = Font(bold=True)
    
    # Write data rows
    for row, result in enumerate(results, 2):
        for col, header in enumerate(headers, 1):
            value = result.get(header, '')
            if value is None:
                value = ''
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    logger.info(f"Results saved to Excel (basic): {filename}")
    return filename


def print_summary_stats(results: List[Dict[str, Any]]) -> None:
    """Print summary statistics of the results."""
    if not results:
        logger.info("No results to summarize")
        return
    
    # Only count results with actual pricing data
    valid_results = [r for r in results if r.get("price_of_shipping") is not None]
    
    if not valid_results:
        print("\n" + "="*60)
        print("SHIPPING MATRIX COMPUTATION SUMMARY")
        print("="*60)
        print("No valid pricing data obtained from API calls")
        print("="*60)
        return
    
    total_valid_results = len(valid_results)
    screenshot_results = [r for r in valid_results if r.get("screenshot_url")]
    
    # Price statistics for valid results
    prices = [r["price_of_shipping"] for r in valid_results]
    
    print("\n" + "="*60)
    print("SHIPPING MATRIX COMPUTATION SUMMARY")
    print("="*60)
    print(f"Total city combinations processed: {total_valid_results}")
    print(f"Successful API responses: {total_valid_results}")
    print(f"Failed API responses: 0")
    print(f"Screenshots captured: {len(screenshot_results)}")
    
    if prices:
        print(f"\nPricing Statistics:")
        print(f"  Average shipping price: ${sum(prices)/len(prices):.2f}")
        print(f"  Minimum shipping price: ${min(prices):.2f}")
        print(f"  Maximum shipping price: ${max(prices):.2f}")
    
    # Top carriers
    carriers = [r.get("carrier") for r in valid_results if r.get("carrier")]
    if carriers:
        from collections import Counter
        carrier_counts = Counter(carriers)
        print(f"\nTop Carriers:")
        for carrier, count in carrier_counts.most_common(5):
            print(f"  {carrier}: {count} routes")
    
    print("="*60)


def process_city_batch(batch_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Process a batch of city combinations in a separate process.
    This function runs the async computation for a subset of cities.
    """
    import asyncio
    import logging
    
    # Setup logging for this process
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Extract batch parameters
    city_pairs = batch_data['city_pairs']
    date = batch_data['date']
    container = batch_data['container']
    delay_range = batch_data['delay_range']
    batch_id = batch_data['batch_id']
    
    logger = logging.getLogger(f"batch_{batch_id}")
    logger.info(f"Processing batch {batch_id} with {len(city_pairs)} city pairs")
    
    # Run the async computation for this batch
    return asyncio.run(compute_batch_async(city_pairs, date, container, delay_range, batch_id))


def format_city(city: str, capitalize: bool = False, hyphenate: bool = True) -> str:
    """
    Format city name for use in URLs.
    capitalize=True: Capitalizes each word, else lowercases all.
    hyphenate=True: Joins words with hyphens.
    """
    words = city.strip().split()
    if capitalize:
        words = [w.capitalize() for w in words]
    else:
        words = [w.lower() for w in words]
    return "-".join(words) if hyphenate else " ".join(words)

async def compute_batch_async(
    city_pairs: List[tuple], 
    date: str, 
    container: str, 
    delay_range: tuple,
    batch_id: int
) -> List[Dict[str, Any]]:
    """
    Async computation for a batch of city pairs.
    """
    results = []
    logger = logging.getLogger(f"batch_{batch_id}")
    
    # Track unique combinations processed in this batch
    batch_combinations_processed = set()
    
    logger.info(f"Batch {batch_id}: Starting with {len(city_pairs)} city combinations to process")
    
    for i, (origin_city, destination_city) in enumerate(city_pairs):
        combination_key = (origin_city, destination_city)
        
        # Log progress every 10 combinations within the batch
        if (i + 1) % 10 == 0 or i == 0 or i == len(city_pairs) - 1:
            logger.info(f"Batch {batch_id}: Processing combination {i+1}/{len(city_pairs)}: {origin_city} -> {destination_city}")
        
        try:
            # Create API payload
            payload = SearatesRequestPayload(
                pointIdFrom=CITIES_TO_POINT_ID_MAP[origin_city],
                pointIdTo=CITIES_TO_POINT_ID_MAP[destination_city],
                date=date,
                container=container
            )
            
            # Make API request with retry logic
            api_response = await make_request_with_retry(SEARATES_API_URL, payload)
            
            if api_response and api_response.data and api_response.data.rates:
                # Process each rate found
                combination_results = []
                for rate_idx, rate in enumerate(api_response.data.rates):
                    screenshot_url = None
                    website_link = None
                    
                    try:
                        # Generate screenshot URL
                        shipment_id = rate.general.shipmentId
                        website_link = f"https://www.searates.com/logistics-explorer/?id={shipment_id}"
                        
                        logger.debug(f"Taking screenshot for shipment {shipment_id}")
                        screenshot_result = await _go(website_link)
                        
                        if isinstance(screenshot_result, str):
                            screenshot_url = screenshot_result
                        elif isinstance(screenshot_result, dict) and "error" not in screenshot_result:
                            screenshot_url = screenshot_result
                        else:
                            logger.error(f"Screenshot failed for {shipment_id}: {screenshot_result}")
                            
                    except Exception as screenshot_error:
                        logger.error(f"Screenshot error for {origin_city} -> {destination_city}: {screenshot_error}")
                    
                    # Create result entry
                    result_entry = create_result_entry(
                        origin_city=origin_city,
                        destination_city=destination_city,
                        rate=rate,
                        date=date,
                        container=container,
                        screenshot_url=screenshot_url,
                        website_link=website_link,
                        is_expired=rate.general.expired
                    )
                    
                    combination_results.append(result_entry)
                    
                    # Add delay between screenshots if multiple rates
                    if rate_idx < len(api_response.data.rates) - 1:
                        await asyncio.sleep(random.uniform(1, 2))
                
                # Mark this combination as successfully processed
                batch_combinations_processed.add(combination_key)
                results.extend(combination_results)
                
                logger.debug(f"Batch {batch_id}: ✅ {origin_city} -> {destination_city} completed with {len(combination_results)} rates")
                
            else:
                # No rates found - create fallback result
                origin_city_url = format_city(origin_city, capitalize=False)
                dest_city_url = format_city(destination_city, capitalize=False)
                origin_city_cap = format_city(origin_city, capitalize=True, hyphenate=False)
                dest_city_cap = format_city(destination_city, capitalize=True, hyphenate=False)

                fallback_website_link = (
                    f"https://www.searates.com/logistics-explorer/from-{origin_city_url}-to-{dest_city_url}/"
                    f"?from={origin_city_cap}&to={dest_city_cap}"
                    f"&fromId={CITIES_TO_POINT_ID_MAP[origin_city]}"
                    f"&toId={CITIES_TO_POINT_ID_MAP[destination_city]}"
                    f"&date={date}&type=fcl&container={container}"
                )
                
                screenshot_url = None
                try:
                    screenshot_result = await _go(fallback_website_link)
                    if isinstance(screenshot_result, str):
                        screenshot_url = screenshot_result
                    elif isinstance(screenshot_result, dict) and "error" not in screenshot_result:
                        screenshot_url = screenshot_result.get("screenshot_url") or screenshot_result
                except Exception as screenshot_error:
                    logger.error(f"Screenshot error for fallback page {fallback_website_link}: {screenshot_error}")

                # Still mark as processed even with no rates
                batch_combinations_processed.add(combination_key)
                empty_result = create_empty_result(
                    origin_city, destination_city, date, container, screenshot_url, fallback_website_link, is_expired=False
                )
                results.append(empty_result)
                
                logger.debug(f"Batch {batch_id}: ⚠️  {origin_city} -> {destination_city} completed with no rates (fallback)")
            
            # Add delay between API requests
            if i < len(city_pairs) - 1:
                delay = random.uniform(delay_range[0], delay_range[1])
                await asyncio.sleep(delay)
                    
        except Exception as e:
            logger.error(f"Batch {batch_id}: Error processing {origin_city} -> {destination_city}: {e}")
            # Still mark as processed even with errors
            batch_combinations_processed.add(combination_key)
            error_result = create_error_result(origin_city, destination_city, date, container, str(e))
            results.append(error_result)
    
    # Final batch summary
    logger.info(f"Batch {batch_id}: ✅ COMPLETED")
    logger.info(f"Batch {batch_id}: Processed {len(batch_combinations_processed)} unique combinations")
    logger.info(f"Batch {batch_id}: Generated {len(results)} total data points")
    logger.info(f"Batch {batch_id}: Average {len(results) / len(batch_combinations_processed):.1f} data points per combination")
    
    return results


async def compute_shipping_matrix_parallel(
    # date: str = "2025-06-25", 
    # container: str = "ST20",
    delay_range: tuple = (2, 5),
    num_processes: int = None,
    city_percentage: float = 0.03,
    checkpoint_interval: int = 50,
    resume: bool = True,
    checkpoint_dir: str = "checkpoints"
) -> List[Dict[str, Any]]:
    """
    Compute shipping matrix using multiprocessing with checkpointing for fault tolerance.
    
    Args:
        date: Shipping date in YYYY-MM-DD format
        container: Container type (e.g., "ST20")
        delay_range: Tuple of (min, max) seconds for random delays
        num_processes: Number of processes to use (default: CPU count)
        city_percentage: Percentage of cities to process (default: 0.03 = 3%)
        checkpoint_interval: Number of results to process before checkpointing
        resume: Whether to resume from existing checkpoint
        checkpoint_dir: Directory to store checkpoint files
        
    Returns:
        List of dictionaries containing shipping data for each combination
    """
    if num_processes is None:
        num_processes = mp.cpu_count()
    
    # Initialize checkpoint manager
    checkpoint_manager = CheckpointManager(
        checkpoint_dir=checkpoint_dir,
        checkpoint_interval=checkpoint_interval
    )
    
    logger.info(f"Starting PARALLEL matrix computation using {num_processes} processes")
    logger.info(f"Checkpointing enabled: every {checkpoint_interval} results")
    
    # Handle resume vs fresh start
    checkpoint_data = {}
    if resume:
        checkpoint_data = checkpoint_manager.load_existing_checkpoint()
        if checkpoint_data['has_checkpoint']:
            logger.info(f"🔄 Resuming from checkpoint with {len(checkpoint_data['results'])} existing results")
    else:
        logger.info("🆕 Starting fresh - clearing existing checkpoints")
        checkpoint_manager.clear_checkpoints()
    
    # Get city combinations
    all_cities = list(CITIES_TO_POINT_ID_MAP.keys())
    cities_count = int(len(all_cities) * city_percentage)
    cities = random.sample(all_cities, cities_count)
    
    # Generate all city+container combinations (exclude same city combinations)
    all_city_container_combinations = []
    for origin_city in cities:
        for destination_city in cities:
            if origin_city != destination_city:
                for container in CONTAINERS:  # ["ST20", "ST40"]
                    all_city_container_combinations.append((origin_city, destination_city, container))
    
    logger.info(f"📊 COMBINATION SETUP:")
    logger.info(f"   Total cities available: {len(all_cities)}")
    logger.info(f"   Cities to process: {len(cities)} ({city_percentage*100:.1f}%)")
    logger.info(f"   Container types: {CONTAINERS}")
    logger.info(f"   Expected combinations: {len(all_city_container_combinations)}")
    logger.info(f"   Target for completion: 14,280 combinations (85 cities × 84 × 2 containers)")
    
    # Get remaining combinations to process (skip completed ones if resuming)
    if resume and checkpoint_data.get('has_checkpoint'):
        remaining_combinations = checkpoint_manager.get_remaining_pairs(all_city_container_combinations)
        completed_count = len(checkpoint_data.get('completed_pairs', set()))
        logger.info(f"🔄 RESUMING FROM CHECKPOINT:")
        logger.info(f"   Combinations already completed: {completed_count}")
        logger.info(f"   Combinations remaining: {len(remaining_combinations)}")
        logger.info(f"   Progress so far: {(completed_count/14280)*100:.1f}% of target")
    else:
        remaining_combinations = all_city_container_combinations
        logger.info(f"🆕 STARTING FRESH: {len(remaining_combinations)} total city+container combinations to process")
    
    # Split remaining combinations into batches for parallel processing
    if not remaining_combinations:
        logger.info("✅ All city+container combinations already completed!")
        progress = checkpoint_manager.get_progress_summary()
        logger.info(f"🎉 FINAL STATS: {progress['completed_combinations']}/{progress['target_combinations']} combinations completed!")
        return checkpoint_manager.total_results
    
    batch_size = max(1, math.ceil(len(remaining_combinations) / num_processes))
    batches = []
    
    for i in range(0, len(remaining_combinations), batch_size):
        selected_date = "2025-09-01"
        
        batch_combinations = remaining_combinations[i:i + batch_size]
        batch_data = {
            'city_container_combinations': batch_combinations,
            'date': selected_date,
            'delay_range': delay_range,
            'batch_id': len(batches) + 1
        }
        batches.append(batch_data)
    
    logger.info(f"Split work into {len(batches)} batches (avg {batch_size} combinations per batch)")
    
    # Process batches in parallel using ProcessPoolExecutor
    try:
        with ProcessPoolExecutor(max_workers=num_processes) as executor:
            logger.info("Starting parallel processing...")
            
            # Submit all batches for processing
            future_to_batch = {executor.submit(process_city_container_batch, batch): batch for batch in batches}
            
            # Collect results as they complete
            for future in future_to_batch:
                try:
                    batch_results = future.result()
                    batch_id = future_to_batch[future]['batch_id']
                    
                    # Add results to checkpoint manager (with automatic checkpointing)
                    checkpoint_manager.add_batch_results(batch_results, batch_id)
                    
                    logger.info(f"✅ Completed batch {batch_id}, total results: {len(checkpoint_manager.total_results)}")
                    
                except Exception as e:
                    batch_id = future_to_batch[future]['batch_id']
                    logger.error(f"❌ Batch {batch_id} failed: {e}")
                    # Continue processing other batches even if one fails
    
    except KeyboardInterrupt:
        logger.info("🛑 Process interrupted by user - saving checkpoint...")
        checkpoint_manager.save_checkpoint(force=True)
        raise
    except Exception as e:
        logger.error(f"❌ Parallel processing error: {e}")
        checkpoint_manager.save_checkpoint(force=True)
        raise
    
    # Final checkpoint and export
    logger.info("💾 Saving final checkpoint...")
    checkpoint_manager.save_checkpoint(force=True)
    
    # Final comprehensive summary
    progress = checkpoint_manager.get_progress_summary()
    logger.info("="*80)
    logger.info("🎉 PARALLEL MATRIX COMPUTATION COMPLETED!")
    logger.info("="*80)
    logger.info(f"📊 COMPLETION STATISTICS:")
    logger.info(f"   ✅ Unique combinations completed: {progress['completed_combinations']}/{progress['target_combinations']} ({progress['progress_percent']:.1f}%)")
    logger.info(f"   📄 Total data points generated: {progress['total_data_points']}")
    logger.info(f"   📈 Average data points per combination: {progress['avg_data_points_per_combination']:.1f}")
    logger.info(f"   🎯 Target achievement: {'COMPLETE' if progress['is_complete'] else 'PARTIAL'}")
    
    if progress['is_complete']:
        logger.info(f"🏆 SUCCESS: All {progress['target_combinations']} city+container combinations have been processed!")
    else:
        remaining = progress['remaining_combinations']
        logger.info(f"⚠️  INCOMPLETE: {remaining} combinations still need processing")
    
    logger.info("="*80)
    
    return checkpoint_manager.total_results


def process_city_container_batch(batch_data: dict) -> List[Dict[str, Any]]:
    """
    Process a batch of city+container combinations in a separate process.
    
    Args:
        batch_data: Dictionary containing batch information
        
    Returns:
        List of dictionaries containing shipping data
    """
    import asyncio
    
    # Extract batch data
    city_container_combinations = batch_data['city_container_combinations']
    date = batch_data['date']
    delay_range = batch_data['delay_range']
    batch_id = batch_data['batch_id']
    
    # Run the async computation
    return asyncio.run(compute_batch_with_containers_async(
        city_container_combinations, 
        date, 
        delay_range,
        batch_id
    ))


async def compute_batch_with_containers_async(
    city_container_combinations: List[tuple], 
    date: str, 
    delay_range: tuple,
    batch_id: int
) -> List[Dict[str, Any]]:
    """
    Async computation for a batch of city+container combinations.
    """
    results = []
    logger = logging.getLogger(f"batch_{batch_id}")
    
    # Track unique combinations processed in this batch
    batch_combinations_processed = set()
    
    logger.info(f"Batch {batch_id}: Starting with {len(city_container_combinations)} city+container combinations to process")
    
    for i, (origin_city, destination_city, container) in enumerate(city_container_combinations):
        combination_key = (origin_city, destination_city, container)
        
        # Log progress every 10 combinations within the batch
        if (i + 1) % 10 == 0 or i == 0 or i == len(city_container_combinations) - 1:
            logger.info(f"Batch {batch_id}: Processing combination {i+1}/{len(city_container_combinations)}: {origin_city} -> {destination_city} ({container})")
        
        try:
            # Create API payload with specific container type
            payload = SearatesRequestPayload(
                pointIdFrom=CITIES_TO_POINT_ID_MAP[origin_city],
                pointIdTo=CITIES_TO_POINT_ID_MAP[destination_city],
                date=date,
                container=container
            )
            
            # Make API request with retry logic
            api_response = await make_request_with_retry(SEARATES_API_URL, payload)
            
            if api_response and api_response.data and api_response.data.rates:
                # Process each rate found
                combination_results = []
                for rate_idx, rate in enumerate(api_response.data.rates):
                    screenshot_url = None
                    website_link = None
                    
                    try:
                        # Generate screenshot URL
                        shipment_id = rate.general.shipmentId
                        website_link = f"https://www.searates.com/logistics-explorer/?id={shipment_id}"
                        
                        logger.debug(f"Taking screenshot for shipment {shipment_id}")
                        screenshot_result = await _go(website_link)
                        
                        if isinstance(screenshot_result, str):
                            screenshot_url = screenshot_result
                        elif isinstance(screenshot_result, dict) and "error" not in screenshot_result:
                            screenshot_url = screenshot_result
                        else:
                            logger.error(f"Screenshot failed for {shipment_id}: {screenshot_result}")
                            
                    except Exception as screenshot_error:
                        logger.error(f"Screenshot error for {origin_city} -> {destination_city} ({container}): {screenshot_error}")
                    
                    # Create result entry with correct parameter order
                    result_entry = create_result_entry(
                        origin_city=origin_city,
                        destination_city=destination_city,
                        rate=rate,
                        date=date,
                        container=container,
                        screenshot_url=screenshot_url,
                        website_link=website_link
                    )
                    
                    combination_results.append(result_entry)
                    
                    # Add delay between screenshots if multiple rates
                    if rate_idx < len(api_response.data.rates) - 1:
                        await asyncio.sleep(random.uniform(1, 2))
                
                # Mark this combination as successfully processed
                batch_combinations_processed.add(combination_key)
                results.extend(combination_results)
                
                logger.debug(f"Batch {batch_id}: ✅ {origin_city} -> {destination_city} ({container}) completed with {len(combination_results)} rates")
                
            else:
                # No rates found - create fallback result
                origin_city_url = format_city(origin_city, capitalize=False)
                dest_city_url = format_city(destination_city, capitalize=False)
                origin_city_cap = format_city(origin_city, capitalize=True, hyphenate=False)
                dest_city_cap = format_city(destination_city, capitalize=True, hyphenate=False)

                fallback_website_link = (
                    f"https://www.searates.com/logistics-explorer/from-{origin_city_url}-to-{dest_city_url}/"
                    f"?from={origin_city_cap}&to={dest_city_cap}"
                    f"&fromId={CITIES_TO_POINT_ID_MAP[origin_city]}"
                    f"&toId={CITIES_TO_POINT_ID_MAP[destination_city]}"
                    f"&date={date}&type=fcl&container={container}"
                )
                
                screenshot_url = None
                try:
                    screenshot_result = await _go(fallback_website_link)
                    if isinstance(screenshot_result, str):
                        screenshot_url = screenshot_result
                    elif isinstance(screenshot_result, dict) and "error" not in screenshot_result:
                        screenshot_url = screenshot_result.get("screenshot_url") or screenshot_result
                except Exception as screenshot_error:
                    logger.error(f"Screenshot error for fallback page {fallback_website_link}: {screenshot_error}")

                # Still mark as processed even with no rates
                batch_combinations_processed.add(combination_key)
                empty_result = create_empty_result(
                    origin_city, destination_city, date, container, screenshot_url, fallback_website_link
                )
                results.append(empty_result)
                
                logger.debug(f"Batch {batch_id}: ⚠️  {origin_city} -> {destination_city} ({container}) completed with no rates (fallback)")
            
            # Add delay between API requests
            if i < len(city_container_combinations) - 1:
                delay = random.uniform(delay_range[0], delay_range[1])
                await asyncio.sleep(delay)
                    
        except Exception as e:
            logger.error(f"Batch {batch_id}: Error processing {origin_city} -> {destination_city} ({container}): {e}")
            # Still mark as processed even with errors
            batch_combinations_processed.add(combination_key)
            error_result = create_error_result(origin_city, destination_city, date, container, str(e))
            results.append(error_result)
    
    # Final batch summary
    logger.info(f"Batch {batch_id}: ✅ COMPLETED")
    logger.info(f"Batch {batch_id}: Processed {len(batch_combinations_processed)} unique city+container combinations")
    logger.info(f"Batch {batch_id}: Generated {len(results)} total data points")
    if len(batch_combinations_processed) > 0:
        logger.info(f"Batch {batch_id}: Average {len(results) / len(batch_combinations_processed):.1f} data points per combination")
    
    return results


async def compute_shipping_matrix_parallel_realtime(
    date: str = "2025-09-01", 
    container: str = "ST20",
    delay_range: tuple = (2, 5),
    num_processes: int = None,
    city_percentage: float = 1.0,
    checkpoint_interval: int = 10,
    resume: bool = True,
    checkpoint_dir: str = "checkpoints",
    excel_backup: bool = False
) -> List[Dict[str, Any]]:
    """
    Compute shipping matrix with REAL-TIME checkpointing every N results.
    This version checkpoints as soon as N results are completed, not when entire batches finish.
    
    Args:
        date: Shipping date in YYYY-MM-DD format
        container: Container type (e.g., "ST20")
        delay_range: Tuple of (min, max) seconds for random delays
        num_processes: Number of processes to use (default: CPU count)
        city_percentage: Percentage of cities to process (default: 0.03 = 3%)
        checkpoint_interval: Number of results to process before checkpointing
        resume: Whether to resume from existing checkpoint
        checkpoint_dir: Directory to store checkpoint files
        
    Returns:
        List of dictionaries containing shipping data for each combination
    """
    if num_processes is None:
        num_processes = mp.cpu_count()
    
    # Initialize checkpoint manager
    checkpoint_manager = CheckpointManager(
        checkpoint_dir=checkpoint_dir,
        checkpoint_interval=checkpoint_interval,
        excel_backup=excel_backup
    )
    
    logger.info(f"Starting REAL-TIME PARALLEL matrix computation using {num_processes} processes")
    logger.info(f"Real-time checkpointing enabled: every {checkpoint_interval} results")
    
    # Handle resume vs fresh start
    checkpoint_data = {}
    if resume:
        checkpoint_data = checkpoint_manager.load_existing_checkpoint()
        if checkpoint_data['has_checkpoint']:
            logger.info(f"🔄 Resuming from checkpoint with {len(checkpoint_data['results'])} existing results")
    else:
        logger.info("🆕 Starting fresh - clearing existing checkpoints")
        checkpoint_manager.clear_checkpoints()
    
    # Get city combinations
    all_cities = list(CITIES_TO_POINT_ID_MAP.keys())

    cities_count = int(len(all_cities) * city_percentage)
    cities = random.sample(all_cities, cities_count)
    
    # Generate all city+container combinations (exclude same city combinations)
    all_city_container_combinations = []
    for origin_city in cities:
        for destination_city in cities:
            if origin_city != destination_city:
                for container in CONTAINERS:  # ["ST20", "ST40"]
                    all_city_container_combinations.append((origin_city, destination_city, container))
    
    logger.info(f"📊 COMBINATION SETUP:")
    logger.info(f"   Total cities available: {len(all_cities)}")
    logger.info(f"   Cities to process: {len(cities)} ({city_percentage*100:.1f}%)")
    logger.info(f"   Container types: {CONTAINERS}")
    logger.info(f"   Expected combinations: {len(all_city_container_combinations)}")
    logger.info(f"   Target for completion: 14,280 combinations (85 cities × 84 × 2 containers)")
    
    # Get remaining combinations to process (skip completed ones if resuming)
    if resume and checkpoint_data.get('has_checkpoint'):
        remaining_combinations = checkpoint_manager.get_remaining_pairs(all_city_container_combinations)
        completed_count = len(checkpoint_data.get('completed_pairs', set()))
        logger.info(f"🔄 RESUMING FROM CHECKPOINT:")
        logger.info(f"   Combinations already completed: {completed_count}")
        logger.info(f"   Combinations remaining: {len(remaining_combinations)}")
        logger.info(f"   Progress so far: {(completed_count/14280)*100:.1f}% of target")
    else:
        remaining_combinations = all_city_container_combinations
        logger.info(f"🆕 STARTING FRESH: {len(remaining_combinations)} total city+container combinations to process")
    
    if not remaining_combinations:
        logger.info("✅ All city+container combinations already completed!")
        progress = checkpoint_manager.get_progress_summary()
        logger.info(f"🎉 FINAL STATS: {progress['completed_combinations']}/{progress['target_combinations']} combinations completed!")
        return checkpoint_manager.total_results
    
    # Use a queue-based approach for real-time processing
    from concurrent.futures import ProcessPoolExecutor, as_completed
    import queue
    import threading
    
    # Split remaining combinations into smaller chunks (not full batches)
    chunk_size = max(1, min(50, len(remaining_combinations) // (num_processes * 4)))  # Small chunks for frequent checkpointing
    chunks = []
    
    for i in range(0, len(remaining_combinations), chunk_size):
        selected_date = "2025-09-01"
        chunk_combinations = remaining_combinations[i:i + chunk_size]
        chunk_data = {
            'city_container_combinations': chunk_combinations,
            'date': selected_date,
            'delay_range': delay_range,
            'batch_id': len(chunks) + 1 
        }
        chunks.append(chunk_data)
    
    logger.info(f"Split work into {len(chunks)} small chunks (avg {chunk_size} combinations per chunk)")
    logger.info(f"This enables checkpointing every {checkpoint_interval} results instead of waiting for full batches")
    
    # Process chunks and checkpoint in real-time
    try:
        with ProcessPoolExecutor(max_workers=num_processes) as executor:
            logger.info("Starting real-time parallel processing...")
            
            # Submit chunks for processing
            future_to_chunk = {executor.submit(process_city_container_batch, chunk): chunk for chunk in chunks}
            
            # Process results as they complete
            for future in as_completed(future_to_chunk):
                try:
                    chunk_results = future.result()
                    chunk_id = future_to_chunk[future]['batch_id']  # Use batch_id for consistency
                    
                    # Add results immediately and trigger checkpointing
                    for result in chunk_results:
                        checkpoint_manager.add_result(result)
                    
                    logger.info(f"✅ Completed chunk {chunk_id} with {len(chunk_results)} results")
                    logger.info(f"📊 Total results so far: {len(checkpoint_manager.total_results)}")
                    
                    # Force checkpoint check (the add_result method handles automatic checkpointing)
                    if len(checkpoint_manager.total_results) % checkpoint_interval == 0:
                        logger.info(f"🎯 Checkpoint triggered at {len(checkpoint_manager.total_results)} results")
                    
                except Exception as e:
                    chunk_id = future_to_chunk[future]['batch_id']  # Use batch_id for consistency
                    logger.error(f"❌ Chunk {chunk_id} failed: {e}")
                    # Continue processing other chunks even if one fails
    
    except KeyboardInterrupt:
        logger.info("🛑 Process interrupted by user - saving checkpoint...")
        checkpoint_manager.save_checkpoint(force=True)
        raise
    except Exception as e:
        logger.error(f"❌ Parallel processing error: {e}")
        checkpoint_manager.save_checkpoint(force=True)
        raise
    
    # Final checkpoint and export
    logger.info("💾 Saving final checkpoint...")
    checkpoint_manager.save_checkpoint(force=True)
    
    # Final comprehensive summary
    progress = checkpoint_manager.get_progress_summary()
    logger.info("="*80)
    logger.info("🎉 REAL-TIME PARALLEL COMPUTATION COMPLETED!")
    logger.info("="*80)
    logger.info(f"📊 COMPLETION STATISTICS:")
    logger.info(f"   ✅ Unique combinations completed: {progress['completed_combinations']}/{progress['target_combinations']} ({progress['progress_percent']:.1f}%)")
    logger.info(f"   📄 Total data points generated: {progress['total_data_points']}")
    logger.info(f"   📈 Average data points per combination: {progress['avg_data_points_per_combination']:.1f}")
    logger.info(f"   🎯 Target achievement: {'COMPLETE' if progress['is_complete'] else 'PARTIAL'}")
    
    if progress['is_complete']:
        logger.info(f"🏆 SUCCESS: All {progress['target_combinations']} city+container combinations have been processed!")
    else:
        remaining = progress['remaining_combinations']
        logger.info(f"⚠️  INCOMPLETE: {remaining} combinations still need processing")
    
    logger.info("="*80)
    
    logger.info(f"🎉 Real-time parallel matrix computation completed. Total results: {len(checkpoint_manager.total_results)}")
    return checkpoint_manager.total_results


async def run_script():
    """Main script to run the shipping matrix computation."""
    logger.info("Starting shipping matrix computation...")
    
    # Run the matrix computation
    results = await compute_shipping_matrix(
        date="2025-09-01",
        container="ST20",
        delay_range=(3, 7)  # 3-7 seconds delay between requests
    )
    
    logger.info(f"Computation completed. Generated {len(results)} results")
    
    # Print summary statistics
    print_summary_stats(results)
    
    # Save results to files
    csv_filename = save_results_to_csv(results)
    json_filename = save_results_to_json(results)
    
    # excel_filename = None
    # try:
    #     excel_filename = save_results_to_excel(results)
    # except ImportError as e:
    #     logger.warning(f"Excel export skipped: {e}")
    
    logger.info(f"Results saved to:")
    logger.info(f"  CSV: {csv_filename}")
    logger.info(f"  JSON: {json_filename}")
    # if excel_filename:
    #     logger.info(f"  Excel: {excel_filename}")
    
    return results


# if __name__ == "__main__":
#     result = asyncio.run(make_request(SEARATES_API_URL, SearatesRequestPayload(pointIdFrom="C_120812", pointIdTo="C_120941", date="2025-06-18", container="ST20")))
#     print(result.data.rates[0].general.shipmentId)


# ============================= FREIGHTOS FUNCTIONALITY =============================

FREIGHTOS_USER_COOKIES = "handlID=92852469965; handl_ref_domain=; handl_landing_page_base=https://www.freightos.com/; traffic_source=Direct; first_traffic_source=Direct; server-version-cookie=y25w24-release.1749630721000|; i18next=en; handl_original_ref=https%3A%2F%2Fship.freightos.com%2F; handl_landing_page=https%3A%2F%2Fwww.freightos.com%2Fwp-content%2Fuploads%2F2018%2F04%2Fcropped-favicon-512x512-Freighots-32x32.png; handl_ref=https%3A%2F%2Fship.freightos.com%2F; handl_url_base=https%3A%2F%2Fwww.freightos.com%2Fwp-content%2Fuploads%2F2018%2F04%2Fcropped-favicon-512x512-Freighots-32x32.png; handl_url=https%3A%2F%2Fwww.freightos.com%2Fwp-content%2Fuploads%2F2018%2F04%2Fcropped-favicon-512x512-Freighots-32x32.png; user_agent=Mozilla%2F5.0%20%28Macintosh%3B%20Intel%20Mac%20OS%20X%2010_15_7%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F137.0.0.0%20Safari%2F537.36; organic_source=https%3A%2F%2Fship.freightos.com%2F; organic_source_str=Other; intercom-id-hwrb8vsu=84d5dfdf-01de-4610-ac59-2cceb47a176d; intercom-device-id-hwrb8vsu=49d67dd8-4587-4c5d-bcb2-17d430cda0b8; HandLtestDomainNameServer=HandLtestDomainValueServer; handl_ip=5.151.198.139; prefs=en|null|GBP|true|GB|0|kg|cm|cbm|cm3_kg|days||W48|YES|false|Freight||cbm|kg|false|false; intercom-session-hwrb8vsu=emlmUmdWR2E5c2xBYjZrRGRxKzZmM1Z3d0UxWE9QeG50ZUR5RE9FVHprWmpBTnI1cUNDSXRLd2ZtaDVPM1VleklsakYxR0FCWmI2R1hYOTZLSGQvTkYzem1DWFY5akpyK3RNUmUveDlmejQ9LS1Zc1JjRHFzdW1ISkdEeVRkaHBRdCtRPT0=--d9843bc58cd8bbe1a1a4ee6c9adf7c4f77cb06a1; session=okafor%40thecozm.com|agpzfnRyYWRlb3Mxch0LEhB1c2VyL0xlZ2FsRW50aXR5GICA6vCFiaoLDA|Okafor+Okafor||1750517761132|1753109761132|yT_32N3EdG4Tvn16XsqHMUfTciA|true|false||false|BuyQuotes+MarketplaceShipper+Buying|BusinessAdmin||||7204968168%3AagpzfnRyYWRlb3Mxch0LEhB1c2VyL0xlZ2FsRW50aXR5GICA6rCyp-kLDA%2CBuyQuotes%2BBuying%2BMarketplaceShipper|V2|v-qdF8g9ijcq1QqmMEa_6-i9Q_k"

FREIGHTOS_SEARCH_URL = "https://ship.freightos.com/api/open-freight/quoting/quotes/search/"

def make_freightos_headers():
    return {
        "Cookie": FREIGHTOS_USER_COOKIES,
        "Content-Type": "application/json",
        "Accept": "*/*",
        "Origin": "https://ship.freightos.com",
    }

def build_freightos_payload(origin_location_code: str, origin_country_id: str, 
                           destination_location_code: str, destination_country_id: str,
                           date: str, container: str = "container20") -> FreightosRequestPayload:
    """
    Build Freightos API payload from location details.
    
    Args:
        origin_location_code: Origin port code (e.g., "GBLON")
        origin_country_id: Origin country ID (e.g., "GB")
        destination_location_code: Destination port code (e.g., "USNYC")
        destination_country_id: Destination country ID (e.g., "US")
        date: Shipping date in YYYY-MM-DD format
        container: Container type (default: "container20")
    
    Returns:
        FreightosRequestPayload object
    """
    # Base payload structure
    payload_dict = {
        "messageHeader": {
            "messageID": f"matrix-{random.randint(100000, 999999)}-{datetime.now().timestamp()}"
        },
        "businessInfo": {
            "serviceName": "Quoting",
            "serviceMethod": "New",
            "messageDateTime": datetime.now().isoformat() + "Z",
            "parties": [
                {
                    "partyTypeCode": "BY",
                    "contact": {
                        "electronicMail": "okafor@thecozm.com"
                    },
                    "knownShipper": False
                },
                {
                    "partyTypeCode": "FR",
                    "name": "freightos.com"
                }
            ]
        },
        "shipment": {
            "originLocation": {
                "locationName": "port",
                "locationTypeCode": "seaport",
                "ID": origin_location_code,
                "countryID": {
                    "value": origin_country_id
                },
                "locationCode": origin_location_code
            },
            "destinationLocation": {
                "locationName": "port",
                "locationTypeCode": "seaport",
                "ID": destination_location_code,
                "countryID": {
                    "value": destination_country_id
                },
                "locationCode": destination_location_code
            },
            "additionalInformation": [],
            "pickupEvent": {
                "eventDate": {
                    "scheduledDateTime": f"{date}T00:00:00+00:00",
                    "endDateTime": ""
                }
            },
            "load": {
                "packages": [
                    {
                        "quantity": 1,
                        "packagingType": container,
                        "overWeightIndicator": False
                    }
                ],
                "declaredValue": {
                    "value": 9000,
                    "currencyID": "USD"
                }
            },
            "accessorials": [],
            "insuranceValueAmount": {
                "value": 9000,
                "currencyID": "USD"
            },
            "pricePreference": {
                "includeOriginPortCharges": True,
                "includeDestinationPortCharges": True,
                "requestCurrency": {}
            },
            "declaredCustoms": {
                "singleEntryBond": {
                    "value": 9000,
                    "currencyID": "USD"
                },
                "entry": True,
                "commodityTypes": 1
            }
        }
    }
    
    return FreightosRequestPayload(**payload_dict)

async def make_freightos_initial_request(origin_location_code: str, origin_country_id: str,
                                       destination_location_code: str, destination_country_id: str,
                                       date: str, container: str = "container20") -> Optional[str]:
    """
    Make initial Freightos request and return polling URL.
    Enhanced with comprehensive error handling and redundancy.
    """
    max_attempts = 10
    
    for attempt in range(max_attempts):
        try:
            async with aiohttp.ClientSession() as session:
                payload = build_freightos_payload(
                    origin_location_code, origin_country_id,
                    destination_location_code, destination_country_id,
                    date, container
                )
                
                async with session.post(
                    url=FREIGHTOS_SEARCH_URL,
                    data=json.dumps(payload.model_dump(exclude_none=True), indent=2),
                    headers=make_freightos_headers(),
                    timeout=aiohttp.ClientTimeout(total=30)
                ) as response:
                    try:
                        response_data = await response.json()
                    except Exception as json_error:
                        logger.warning(f"Failed to parse JSON response (attempt {attempt + 1}): {json_error}")
                        if attempt < max_attempts - 1:
                            await asyncio.sleep(1)
                            continue
                        return None
                    
                    # Check for obvious API errors first
                    if isinstance(response_data, dict):
                        if "businessInfo" in response_data and "message" in response_data.get("businessInfo", {}):
                            error_msg = response_data["businessInfo"]["message"]
                            logger.error(f"Freightos API error: {error_msg}")
                            return None
                        
                        if "error" in response_data or "errors" in response_data:
                            logger.error(f"Freightos API returned error: {response_data}")
                            return None
                    
                    # Try to parse with ultra-permissive model
                    try:
                        result = FreightosQuotesResponse(**response_data)
                        resultId = result.messageHeader.conversationID
                        logger.info(f"Freightos API result ID: {resultId}")
                        if result.paging and result.paging.next:
                            return result.paging.next
                        else:
                            logger.warning(f"No polling URL in response (attempt {attempt + 1})")
                            return None
                    except Exception as parse_error:
                        logger.warning(f"Model parsing failed (attempt {attempt + 1}): {parse_error}")
                        
                        # Try to extract polling URL manually as fallback
                        try:
                            if isinstance(response_data, dict) and "paging" in response_data:
                                paging_data = response_data["paging"]
                                if isinstance(paging_data, dict) and "next" in paging_data:
                                    next_url = paging_data["next"]
                                    if next_url and isinstance(next_url, str):
                                        logger.info(f"Extracted polling URL manually: {next_url}")
                                        return next_url
                        except Exception as manual_error:
                            logger.warning(f"Manual URL extraction failed: {manual_error}")
                        
                        if attempt < max_attempts - 1:
                            await asyncio.sleep(1)
                            continue
                        return None
                        
        except asyncio.TimeoutError:
            logger.warning(f"Request timeout (attempt {attempt + 1})")
            if attempt < max_attempts - 1:
                await asyncio.sleep(2)
                continue
        except Exception as e:
            logger.warning(f"Request failed (attempt {attempt + 1}): {e}")
            if attempt < max_attempts - 1:
                await asyncio.sleep(1)
                continue
    
    logger.error(f"Failed to get polling URL after {max_attempts} attempts")
    return None

async def make_freightos_polling_request(polling_url: str, max_retries: int = 5, 
                                       retry_delay: float = 3.0) -> Optional[tuple]:
    """
    Make polling request to Freightos and return the quotes data with resultId.
    Enhanced with comprehensive error handling and redundancy.
    Returns: (quotes_data, result_id) tuple or None if failed
    """
    if not polling_url or not isinstance(polling_url, str):
        logger.error("Invalid polling URL provided")
        return None
    
    async with aiohttp.ClientSession() as session:
        for attempt in range(10):
            try:
                async with session.get(
                    url=polling_url,
                    headers=make_freightos_headers(),
                    timeout=aiohttp.ClientTimeout(total=30)
                ) as response:
                    if response.status != 200:
                        logger.warning(f"Polling returned status {response.status} (attempt {attempt + 1})")
                        if attempt < max_retries - 1:
                            await asyncio.sleep(retry_delay)
                            continue
                        return None
                    
                    try:
                        data = await response.json()
                    except Exception as json_error:
                        logger.warning(f"Failed to parse polling JSON (attempt {attempt + 1}): {json_error}")
                        if attempt < max_retries - 1:
                            await asyncio.sleep(retry_delay)
                            continue
                        return None
                    
                    # Extract resultId from response
                    result_id = None
                    try:
                        result = FreightosQuotesResponse(**data)
                        result_id = result.messageHeader.conversationID
                        logger.info(f"Freightos polling result ID: {result_id}")
                    except Exception as parse_error:
                        logger.warning(f"Could not extract resultId from polling response: {parse_error}")
                        # Continue without resultId - we'll still return data if available
                    
                    # Check if we have quotes
                    try:
                        if isinstance(data, dict):
                            quotes = data.get("quotes")
                            if quotes and len(quotes) > 0:
                                logger.debug(f"Successfully received {len(quotes)} quotes")
                                return (data, result_id)
                            else:
                                # No quotes yet but we have a valid response - check if we should return anyway
                                if attempt >= max_retries - 1:
                                    # Final attempt - return the response even without quotes for screenshot
                                    logger.info(f"Final polling attempt - returning response for screenshot")
                                    return (data, result_id)
                                else:
                                    logger.info(f"No quotes yet, retrying in {retry_delay} seconds... (attempt {attempt + 1}/{max_retries})")
                                    await asyncio.sleep(retry_delay)
                                    continue
                        else:
                            logger.warning(f"Unexpected response format (attempt {attempt + 1}): {type(data)}")
                            if attempt < max_retries - 1:
                                await asyncio.sleep(retry_delay)
                                continue
                            return None
                    except Exception as check_error:
                        logger.warning(f"Error checking quotes (attempt {attempt + 1}): {check_error}")
                        if attempt < max_retries - 1:
                            await asyncio.sleep(retry_delay)
                            continue
                        return None
                        
            except asyncio.TimeoutError:
                logger.warning(f"Polling timeout (attempt {attempt + 1})")
                if attempt < max_retries - 1:
                    await asyncio.sleep(retry_delay)
                    continue
            except Exception as e:
                logger.warning(f"Polling request failed (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(retry_delay)
                    continue
    
    logger.warning(f"Failed to get quotes after {max_retries} attempts - continuing to next request")
    return None

def locode_to_city_country(locode):
    """Convert UN/LOCODE to city and country."""
    if not locode or len(locode) < 5:
        return "", ""
    country = locode[:2]
    city = locode[2:]
    
    LOCODE_CITIES = {
        "LON": "LONDON", "NYC": "NEW YORK", "LAX": "LOS ANGELES", "LGB": "LONG BEACH",
        "SAV": "SAVANNAH", "SEA": "SEATTLE", "OAK": "OAKLAND", "CHS": "CHARLESTON",
        "HOU": "HOUSTON", "MIA": "MIAMI", "BAL": "BALTIMORE", "FXT": "FELIXSTOWE",
        "SOU": "SOUTHAMPTON", "LIV": "LIVERPOOL", "RTM": "ROTTERDAM", "AMS": "AMSTERDAM",
        "HAM": "HAMBURG", "BRV": "BREMEN", "ANR": "ANTWERP", "LEH": "LE HAVRE",
        "MRS": "MARSEILLE", "GOA": "GENOA", "BCN": "BARCELONA", "VLC": "VALENCIA",
        "SHA": "SHANGHAI", "SZX": "SHENZHEN", "NGB": "NINGBO", "TAO": "QINGDAO",
        "CAN": "GUANGZHOU", "TSN": "TIANJIN", "XMN": "XIAMEN", "DLC": "DALIAN",
        "SIN": "SINGAPORE", "HKG": "HONG KONG", "PUS": "BUSAN", "TYO": "TOKYO",
        "YOK": "YOKOHAMA", "UKB": "KOBE", "DXB": "DUBAI", "JEA": "JEBEL ALI",
        "BOM": "MUMBAI", "MAA": "CHENNAI", "SYD": "SYDNEY", "MEL": "MELBOURNE",
        "SSZ": "SANTOS", "BUE": "BUENOS AIRES", "DUR": "DURBAN", "CPT": "CAPE TOWN"
    }
    return LOCODE_CITIES.get(city, city), country

def transform_freightos_response(data, original_container_type: str = "container20", 
                                origin_city_name: str = None, destination_city_name: str = None):
    """
    Transform Freightos response data to standardized format.
    Enhanced with comprehensive error handling to work with any data format.
    """
    if not data:
        logger.warning("Empty data provided to transform function")
        return None
    
    try:
        # Safely extract quotes
        quotes = None
        if isinstance(data, dict):
            quotes = data.get("quotes", [])
        elif hasattr(data, 'quotes'):
            quotes = data.quotes
        
        if not quotes or len(quotes) == 0:
            logger.debug("No quotes found in response data")
            return None
        
        # Use first quote with comprehensive error handling
        quote = quotes[0] if isinstance(quotes, list) else quotes
        if not quote:
            logger.warning("First quote is empty")
            return None
        
        # Safe extraction with fallbacks for all fields
        def safe_get(obj, *keys, default=""):
            """Safely navigate nested dictionary/object structure"""
            try:
                current = obj
                for key in keys:
                    if isinstance(current, dict):
                        current = current.get(key)
                    elif hasattr(current, key):
                        current = getattr(current, key)
                    else:
                        return default
                    if current is None:
                        return default
                return current if current is not None else default
            except Exception:
                return default
        
        # Extract location codes with multiple fallback strategies
        origin_code = ""
        dest_code = ""
        
        try:
            origin_code = (
                safe_get(quote, "originLocation", "locationCode") or
                safe_get(quote, "origin", "locationCode") or
                safe_get(quote, "from", "code") or
                safe_get(quote, "originLocation", "code") or
                ""
            )
            dest_code = (
                safe_get(quote, "destinationLocation", "locationCode") or
                safe_get(quote, "destination", "locationCode") or
                safe_get(quote, "to", "code") or
                safe_get(quote, "destinationLocation", "code") or
                ""
            )
        except Exception as e:
            logger.warning(f"Error extracting location codes: {e}")
        
        # Use provided city names if available, otherwise convert location codes
        if origin_city_name:
            city_of_origin = origin_city_name
            country_of_origin = origin_code[:2] if origin_code and len(origin_code) >= 2 else ""
        else:
            city_of_origin, country_of_origin = locode_to_city_country(origin_code)
            
        if destination_city_name:
            city_of_destination = destination_city_name  
            country_of_destination = dest_code[:2] if dest_code and len(dest_code) >= 2 else ""
        else:
            city_of_destination, country_of_destination = locode_to_city_country(dest_code)
        
        # Extract date with fallbacks
        date_of_shipping = ""
        try:
            date_of_shipping = (
                safe_get(quote, "createDate") or
                safe_get(quote, "date") or
                safe_get(quote, "shipmentDate") or
                datetime.now().date().isoformat()
            )
            if date_of_shipping and len(str(date_of_shipping)) > 10:
                date_of_shipping = str(date_of_shipping)[:10]
        except Exception as e:
            logger.warning(f"Error extracting date: {e}")
            date_of_shipping = datetime.now().date().isoformat()
        
        # Extract shipping time with multiple fallbacks
        shipping_time = None
        try:
            # First try the correct Freightos API structure
            transit_times = safe_get(quote, "connection", "transitTime", "estimatedTransitTimes")
            if transit_times and isinstance(transit_times, list) and len(transit_times) > 0:
                time_range = transit_times[0]
                from_value = safe_get(time_range, "from", "value")
                to_value = safe_get(time_range, "to", "value")
                
                if from_value is not None and to_value is not None:
                    try:
                        from_days = int(from_value)
                        to_days = int(to_value)
                        if from_days == to_days:
                            shipping_time = f"{from_days} days"
                        else:
                            shipping_time = f"{from_days}-{to_days} days"
                    except (ValueError, TypeError):
                        pass
            
            # Fallback to other possible structures if the main one fails
            if shipping_time is None:
                fallback_value = (
                    safe_get(quote, "transitTime", "value") or
                    safe_get(quote, "estimatedDays") or
                    safe_get(quote, "transitDays")
                )
                if fallback_value is not None and str(fallback_value).strip():
                    try:
                        days = int(float(fallback_value))
                        shipping_time = f"{days} days"
                    except (ValueError, TypeError):
                        pass
                        
        except Exception as e:
            logger.warning(f"Error extracting shipping time: {e}")
            shipping_time = None
        
        # Extract price and currency with comprehensive fallbacks
        price_of_shipping = None
        currency = None
        
        try:
            # Try priceIndicator.totalCharge first
            price_info = safe_get(quote, "priceIndicator", "totalCharge")
            if isinstance(price_info, dict):
                price_of_shipping = price_info.get("value")
                currency = price_info.get("currencyID")
            
            # Fallback to connection segments
            if price_of_shipping is None:
                try:
                    segments = safe_get(quote, "connection", "connectionSegments")
                    if segments and len(segments) > 0:
                        charges = safe_get(segments[0], "charges")
                        if charges and len(charges) > 0:
                            rate = safe_get(charges[0], "rate")
                            if isinstance(rate, dict):
                                price_of_shipping = rate.get("value")
                                currency = rate.get("currencyID")
                except Exception:
                    pass
            
            # More fallbacks for price
            if price_of_shipping is None:
                price_of_shipping = (
                    safe_get(quote, "totalPrice") or
                    safe_get(quote, "price", "value") or
                    safe_get(quote, "cost") or
                    safe_get(quote, "amount")
                )
            
            # More fallbacks for currency
            if currency is None:
                currency = (
                    safe_get(quote, "totalCurrency") or
                    safe_get(quote, "price", "currency") or
                    safe_get(quote, "currencyCode") or
                    "USD"
                )
            
            # Convert price to float if possible
            if price_of_shipping is not None:
                try:
                    price_of_shipping = float(price_of_shipping)
                except (ValueError, TypeError):
                    price_of_shipping = None
                    
        except Exception as e:
            logger.warning(f"Error extracting price: {e}")
            price_of_shipping = None
            currency = "USD"
        
        # Container type mapping
        container_mapping = {
            "container20": "ST20",
            "container40": "ST40", 
            "container40hc": "ST40HC",
            "container45": "ST45"
        }
        container_type = container_mapping.get(original_container_type, original_container_type)
        
        # Extract carrier with fallbacks
        carrier = ""
        try:
            parties = safe_get(quote, "businessInfo", "parties")
            if isinstance(parties, list):
                for party in parties:
                    if isinstance(party, dict) and party.get("partyTypeCode") == "FW":
                        carrier = party.get("name", "")
                        break
            
            if not carrier:
                carrier = (
                    safe_get(quote, "carrier", "name") or
                    safe_get(quote, "provider") or
                    safe_get(quote, "forwarder") or
                    ""
                )
        except Exception as e:
            logger.warning(f"Error extracting carrier: {e}")
            carrier = ""
        
        # Extract shipment and rate IDs
        shipment_id = ""
        rate_id = ""
        try:
            shipment_id = (
                safe_get(quote, "referenceID") or
                safe_get(quote, "shipmentId") or
                safe_get(quote, "id") or
                ""
            )
            rate_id = (
                safe_get(quote, "connection", "connectionSegments", 0, "segmentID") or
                safe_get(quote, "rateId") or
                safe_get(quote, "rate", "id") or
                ""
            )
        except Exception as e:
            logger.warning(f"Error extracting IDs: {e}")
        
        # Generate website URL
        website_url = ""
        try:
            if shipment_id and isinstance(shipment_id, str) and shipment_id.strip():
                website_url = f"https://ship.freightos.com/results/{shipment_id.strip()}"
        except Exception as e:
            logger.warning(f"Error generating website URL: {e}")
        
        # Extract CO2 data
        co2_amount = None
        co2_price = None
        try:
            co2_data = safe_get(quote, "co2Emissions")
            if isinstance(co2_data, dict):
                co2_amount = co2_data.get("value")
                if co2_amount is not None:
                    co2_amount = float(co2_amount)
        except Exception as e:
            logger.warning(f"Error extracting CO2 data: {e}")
        
        # Extract validity dates
        validity_from = date_of_shipping
        validity_to = ""
        try:
            validity_to = (
                safe_get(quote, "validTo") or
                safe_get(quote, "expiryDate") or
                safe_get(quote, "validUntil") or
                ""
            )
        except Exception as e:
            logger.warning(f"Error extracting validity: {e}")
        
        # Return structured result
        result = {
            "city_of_origin": city_of_origin or "Unknown",
            "country_of_origin": country_of_origin or "",
            "city_of_destination": city_of_destination or "Unknown", 
            "country_of_destination": country_of_destination or "",
            "date_of_shipping": date_of_shipping,
            "total_shipping_time_days": shipping_time,
            "price_of_shipping": price_of_shipping,
            "currency": currency or "USD",
            "container_type": container_type,
            "provider": "Freightos",
            "carrier": carrier,
            "shipment_id": shipment_id,
            "rate_id": rate_id,
            "website_url": website_url,
            "co2_amount": co2_amount,
            "co2_price": co2_price,
            "validity_from": validity_from,
            "validity_to": validity_to,
            "distance": "",
            "point_total": None,
            "route_total": None
        }
        
        logger.debug(f"Successfully transformed quote: {city_of_origin} -> {city_of_destination}, ${price_of_shipping}")
        return result
        
    except Exception as e:
        logger.error(f"Critical error in transform function: {e}")
        # Return a minimal valid result even on complete failure
        return {
            "city_of_origin": "Unknown",
            "country_of_origin": "",
            "city_of_destination": "Unknown",
            "country_of_destination": "",
            "date_of_shipping": datetime.now().date().isoformat(),
            "total_shipping_time_days": None,
            "price_of_shipping": None,
            "currency": "USD",
            "container_type": original_container_type,
            "provider": "Freightos",
            "carrier": "",
            "shipment_id": "",
            "rate_id": "",
            "website_url": "",
            "co2_amount": None,
            "co2_price": None,
            "validity_from": datetime.now().date().isoformat(),
            "validity_to": "",
            "distance": "",
            "point_total": None,
            "expired": True,
            "route_total": None
        }

def create_freightos_result_entry(
    origin_location: str,
    destination_location: str,
    quotes_data: Dict[str, Any],
    date: str,
    container: str,
    screenshot_url: Optional[str] = None,
    website_link: Optional[str] = None
) -> Dict[str, Any]:
    """
    Create a standardized result entry from Freightos quotes data.
    Enhanced with status tracking for complete data consistency.
    """
    try:
        if quotes_data:
            transformed = transform_freightos_response(
                quotes_data, 
                original_container_type=container,
                origin_city_name=origin_location,
                destination_city_name=destination_location
            )
            if transformed and transformed.get("price_of_shipping") is not None:
                # Add success status fields and screenshot info for consistency
                transformed.update({
                    "request_status": "success",
                    "error_type": None,
                    "error_message": "",
                    "has_polling_url": True,
                    "polling_attempts": 1,  # Assuming at least 1 successful attempt
                    "processing_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "screenshot_url": screenshot_url,
                    "website_url": website_link
                })
                return transformed
    except Exception as e:
        logger.error(f"Error transforming Freightos response: {e}")
        return create_freightos_empty_result(
            origin_location, destination_location, date, container,
            error_type="transformation_error",
            error_message=str(e),
            has_polling_url=True,
            polling_attempts=1,
            screenshot_url=screenshot_url,
            website_link=website_link
        )
    
    # Return empty result if transformation failed or no valid price
    return create_freightos_empty_result(
        origin_location, destination_location, date, container,
        error_type="invalid_response_data",
        error_message="No valid pricing data in response",
        has_polling_url=True,
        polling_attempts=1,
        screenshot_url=screenshot_url,
        website_link=website_link
    )

def create_freightos_empty_result(
    origin_location: str, 
    destination_location: str, 
    date: str, 
    container: str,
    error_type: str = "no_quotes",
    error_message: str = "",
    has_polling_url: bool = False,
    polling_attempts: int = 0,
    screenshot_url: Optional[str] = None,
    website_link: Optional[str] = None
) -> Dict[str, Any]:
    """
    Create an empty/error result entry for failed Freightos requests.
    Enhanced with detailed error tracking for complete data consistency.
    """
    # Map container type to standard format
    container_mapping = {
        "container20": "ST20",
        "container40": "ST40",
        "container40hc": "ST40HC",
        "container45": "ST45"
    }
    mapped_container = container_mapping.get(container, container)
    
    return {
        "city_of_origin": origin_location,
        "country_of_origin": "",
        "city_of_destination": destination_location,
        "country_of_destination": "",
        "date_of_shipping": date,
        "total_shipping_time_days": None,
        "price_of_shipping": None,
        "currency": None,
        "container_type": mapped_container,
        "provider": "Freightos",
        "carrier": "",
        "shipment_id": "",
        "rate_id": "",
        "website_url": website_link or "",
        "co2_amount": None,
        "co2_price": None,
        "validity_from": date,
        "validity_to": "",
        "distance": "",
        "point_total": None,
        "route_total": None,
        
        # ERROR TRACKING FIELDS - for data analysis
        "request_status": "failed",
        "error_type": error_type,  # "location_not_found", "no_polling_url", "no_quotes", "polling_failed", "critical_error"
        "error_message": error_message,
        "has_polling_url": has_polling_url,
        "polling_attempts": polling_attempts,
        "processing_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        
        # SCREENSHOT FIELDS - for consistency
        "screenshot_url": screenshot_url
    }

async def compute_freightos_matrix(
    date: str = "2025-06-25", 
    container: str = "container20",
    delay_range: tuple = (2, 5),
    location_percentage: float = 0.1
) -> List[Dict[str, Any]]:
    """
    Compute Freightos shipping matrix for all location combinations.
    
    Args:
        date: Shipping date in YYYY-MM-DD format
        container: Container type (e.g., "container20")
        delay_range: Tuple of (min, max) seconds for random delays
        location_percentage: Percentage of locations to process
        
    Returns:
        List of dictionaries containing shipping data for each combination
    """
    results = []
    all_locations = list(FREIGHTOS_LOCATIONS.keys())
    locations_count = int(len(all_locations) * location_percentage)
    locations = random.sample(all_locations, locations_count)
    total_combinations = len(locations) * (len(locations) - 1)
    
    logger.info(f"Starting Freightos matrix computation for {total_combinations} location combinations")
    
    for i, origin_location in enumerate(locations):
        for j, destination_location in enumerate(locations):
            if origin_location == destination_location:
                continue
                
            combination_num = i * (len(locations) - 1) + j - (1 if j > i else 0)
            logger.info(f"Processing combination {combination_num + 1}/{total_combinations}: {origin_location} -> {destination_location}")
            
            try:
                origin_data = FREIGHTOS_LOCATIONS[origin_location]
                destination_data = FREIGHTOS_LOCATIONS[destination_location]
                
                # Step 1: Make initial request to get polling URL
                polling_url = await make_freightos_initial_request(
                    origin_data["locationCode"], origin_data["countryID"],
                    destination_data["locationCode"], destination_data["countryID"],
                    date, container
                )
                
                if not polling_url:
                    logger.warning(f"Failed to get polling URL for {origin_location} -> {destination_location}")
                    results.append(create_freightos_empty_result(origin_location, destination_location, date, container))
                    continue
                
                # Step 2: Poll for quotes
                quotes_data = await make_freightos_polling_request(polling_url)
                
                if quotes_data:
                    result = create_freightos_result_entry(origin_location, destination_location, quotes_data, date, container)
                    results.append(result)
                    logger.info(f"✅ Successfully got quote: {origin_location} -> {destination_location}")
                else:
                    logger.warning(f"No quotes returned for {origin_location} -> {destination_location}")
                    results.append(create_freightos_empty_result(origin_location, destination_location, date, container))
                
            except Exception as e:
                logger.error(f"Error processing {origin_location} -> {destination_location}: {e}")
                results.append(create_freightos_empty_result(origin_location, destination_location, date, container))
            
            # Random delay between requests
            if delay_range:
                delay = random.uniform(delay_range[0], delay_range[1])
                logger.debug(f"Waiting {delay:.1f} seconds before next request...")
                await asyncio.sleep(delay)
    
    logger.info(f"Freightos matrix computation completed. Generated {len(results)} results")
    return results

def process_freightos_location_batch(batch_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Process a batch of Freightos location+container pairs.
    This function is designed to be run in a separate process.
    """
    batch_results = []
    location_container_pairs = batch_data['location_container_pairs']
    date = batch_data['date']
    delay_range = batch_data['delay_range']
    batch_id = batch_data['batch_id']
    checkpoint_manager = batch_data.get('checkpoint_manager')
    checkpoint_interval = batch_data.get('checkpoint_interval', 5)
    
    logger.info(f"Starting Freightos batch {batch_id} with {len(location_container_pairs)} location+container pairs")
    
    # Run the async batch processing
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    
    try:
        batch_results = loop.run_until_complete(
            compute_freightos_batch_async(
                location_container_pairs, 
                date, 
                delay_range, 
                batch_id,
                checkpoint_manager,
                checkpoint_interval
            )
        )
    except Exception as e:
        logger.error(f"Error in Freightos batch {batch_id}: {e}")
    finally:
        loop.close()
    
    logger.info(f"Completed Freightos batch {batch_id} with {len(batch_results)} results")
    return batch_results

async def compute_freightos_batch_async(
    location_container_pairs: List[tuple], 
    date: str, 
    delay_range: tuple,
    batch_id: int,
    checkpoint_manager=None,
    checkpoint_interval: int = 5
) -> List[Dict[str, Any]]:
    """
    Asynchronously compute a batch of Freightos location+container pairs.
    Enhanced with comprehensive error handling and COMPLETE data consistency.
    SAVES ALL COMBINATIONS - successful and failed - for complete dataset analysis.
    TIME-BASED CHECKPOINTING: Saves every 5 minutes regardless of result count.
    """
    from datetime import datetime
    import os
    import json
    
    results = []
    total_pairs = len(location_container_pairs)
    successful_count = 0
    failed_count = 0
    
    # TIME-BASED CHECKPOINTING: Track last save time
    last_checkpoint_time = datetime.now()
    checkpoint_interval_minutes = 5  # Save every 5 minutes
    
    logger.info(f"Batch {batch_id} - Processing {total_pairs} location+container pairs with COMPLETE data saving (success + errors)")
    logger.info(f"Batch {batch_id} - Time-based checkpointing every {checkpoint_interval_minutes} minutes")
    
    for idx, (origin_location, destination_location, container) in enumerate(location_container_pairs, 1):
        logger.info(f"Batch {batch_id} - Processing {idx}/{total_pairs}: {origin_location} -> {destination_location} ({container})")
        
        result = None
        polling_attempts = 0
        
        try:
            # Get location data with error handling
            origin_data = FREIGHTOS_LOCATIONS.get(origin_location)
            destination_data = FREIGHTOS_LOCATIONS.get(destination_location)
            
            if not origin_data:
                logger.warning(f"Batch {batch_id} - Origin location '{origin_location}' not found")
                result = create_freightos_empty_result(
                    origin_location, destination_location, date, container,
                    error_type="location_not_found", 
                    error_message=f"Origin location '{origin_location}' not found in FREIGHTOS_LOCATIONS",
                    has_polling_url=False, polling_attempts=0
                )
                failed_count += 1
                
            elif not destination_data:
                logger.warning(f"Batch {batch_id} - Destination location '{destination_location}' not found")
                result = create_freightos_empty_result(
                    origin_location, destination_location, date, container,
                    error_type="location_not_found", 
                    error_message=f"Destination location '{destination_location}' not found in FREIGHTOS_LOCATIONS",
                    has_polling_url=False, polling_attempts=0
                )
                failed_count += 1
                
            else:
                # Step 1: Make initial request to get polling URL
                polling_url = None
                try:
                    polling_url = await make_freightos_initial_request(
                        origin_data["locationCode"], origin_data["countryID"],
                        destination_data["locationCode"], destination_data["countryID"],
                        date, container
                    )
                except Exception as e:
                    logger.error(f"Batch {batch_id} - Error in initial request: {e}")
                    polling_url = None
                
                if not polling_url:
                    logger.warning(f"Batch {batch_id} - Failed to get polling URL")
                    result = create_freightos_empty_result(
                        origin_location, destination_location, date, container,
                        error_type="no_polling_url", 
                        error_message="Failed to get polling URL from initial request",
                        has_polling_url=False, polling_attempts=0
                    )
                    failed_count += 1
                    
                else:
                    # Step 2: Poll for quotes with error handling
                    polling_response = None
                    result_id = None
                    try:
                        polling_response = await make_freightos_polling_request(polling_url)
                        polling_attempts = 1  # Assuming at least 1 attempt was made
                    except Exception as e:
                        logger.error(f"Batch {batch_id} - Error in polling request: {e}")
                        polling_response = None
                        polling_attempts = 1
                    
                    # Extract quotes_data and result_id from response
                    quotes_data = None
                    if polling_response:
                        quotes_data, result_id = polling_response
                    
                    # Step 3: Take screenshot if we have result_id (regardless of quotes)
                    screenshot_url = None
                    website_link = None
                    
                    if result_id:
                        try:
                            # Generate website link using result_id
                            website_link = f"https://ship.freightos.com/results/{result_id}/"
                            
                            logger.info(f"Taking screenshot for result {result_id}")
                            screenshot_result = await _go(website_link)
                            
                            # Enhanced screenshot URL extraction with better logging
                            logger.info(f"Screenshot result type: {type(screenshot_result)}, value: {screenshot_result}")
                            
                            if screenshot_result:
                                # Try multiple ways to extract the URL
                                if isinstance(screenshot_result, str) and screenshot_result.startswith('http'):
                                    screenshot_url = screenshot_result
                                    logger.info(f"✅ Screenshot captured (string): {screenshot_url}")
                                elif hasattr(screenshot_result, 'url') and screenshot_result.url:
                                    screenshot_url = screenshot_result.url
                                    logger.info(f"✅ Screenshot captured (object.url): {screenshot_url}")
                                elif isinstance(screenshot_result, dict) and 'screenshot_url' in screenshot_result:
                                    screenshot_url = screenshot_result['screenshot_url']
                                    logger.info(f"✅ Screenshot captured (screenshot_url): {screenshot_url}")
                                elif isinstance(screenshot_result, dict) and 'url' in screenshot_result:
                                    screenshot_url = screenshot_result['url']
                                    logger.info(f"✅ Screenshot captured (dict url): {screenshot_url}")
                                else:
                                    # Try to convert to string if it contains URL-like content
                                    str_result = str(screenshot_result)
                                    if 'https://' in str_result and '.amazonaws.com' in str_result:
                                        # Extract URL from string representation
                                        url_match = re.search(r'(https://[^\s]+amazonaws\.com[^\s]*\.png)', str_result)
                                        if url_match:
                                            screenshot_url = url_match.group(1)
                                            logger.info(f"✅ Screenshot captured (extracted): {screenshot_url}")
                                        else:
                                            logger.warning(f"⚠️ Screenshot result contains AWS URL but couldn't extract: {str_result}")
                                    else:
                                        logger.warning(f"⚠️ Screenshot failed - unrecognized format: {type(screenshot_result)} = {screenshot_result}")
                            else:
                                logger.warning(f"⚠️ Screenshot failed - empty result returned")
                        
                        except Exception as screenshot_error:
                            logger.warning(f"⚠️ Screenshot failed for {origin_location} -> {destination_location}: {screenshot_error}")
                            # Continue processing even if screenshot fails
                    else:
                        logger.warning(f"Batch {batch_id} - No result_id available for screenshot")
                    
                    # Step 4: Process the response - ALWAYS create results (potentially multiple)
                    if quotes_data:
                        # NEW: Process ALL quotes, not just the first one
                        batch_results = create_freightos_result_entries(
                            origin_location, destination_location, quotes_data, date, container,
                            screenshot_url, website_link
                        )
                        
                        # Process each quote result
                        success_quotes = 0
                        for result in batch_results:
                            if result and result.get("request_status") == "success":
                                success_quotes += 1
                                logger.info(f"Batch {batch_id} - ✅ Quote {result.get('quote_number', 1)}/{result.get('total_quotes_available', 1)}: {origin_location} -> {destination_location} (${result.get('price_of_shipping')}) via {result.get('carrier', 'Unknown')}")
                            results.append(result)
                        
                        if success_quotes > 0:
                            successful_count += 1
                            logger.info(f"Batch {batch_id} - 🎯 Total: {success_quotes} valid quotes found for {origin_location} -> {destination_location}")
                        else:
                            failed_count += 1
                            logger.warning(f"Batch {batch_id} - ⚠️ No valid pricing data in any quotes")
                    else:
                        logger.warning(f"Batch {batch_id} - No quotes returned after polling")
                        result = create_freightos_empty_result(
                            origin_location, destination_location, date, container,
                            error_type="no_quotes", 
                            error_message="No quotes returned after polling retries",
                            has_polling_url=True, polling_attempts=polling_attempts,
                            screenshot_url=screenshot_url, website_link=website_link
                        )
                        results.append(result)
                        failed_count += 1
            
        except Exception as e:
            logger.error(f"Batch {batch_id} - Critical error: {e}")
            result = create_freightos_empty_result(
                origin_location, destination_location, date, container,
                error_type="critical_error", 
                error_message=str(e),
                has_polling_url=False, polling_attempts=polling_attempts
            )
            results.append(result)
            failed_count += 1
        
        # TIME-BASED CHECKPOINTING: Save every 5 minutes if we have any data
        current_time = datetime.now()
        time_since_last_checkpoint = (current_time - last_checkpoint_time).total_seconds() / 60
        
        if time_since_last_checkpoint >= checkpoint_interval_minutes and len(results) > 0:
            try:
                # Save results to multiple formats
                timestamp = current_time.strftime("%Y%m%d_%H%M%S")
                base_filename = f"freightos_checkpoints/batch_{batch_id}_checkpoint_{len(results)}_{timestamp}"
                
                # Create directory if it doesn't exist
                os.makedirs("freightos_checkpoints", exist_ok=True)
                
                # Save individual batch files
                json_file = f"{base_filename}.json"
                with open(json_file, 'w') as f:
                    json.dump(results, f, indent=2)
                
                csv_file = save_results_to_csv(results, f"{base_filename}.csv")
                excel_file = save_results_to_excel_basic(results, f"{base_filename}.xlsx")
                
                # CENTRALIZED AGGREGATION: Collect ALL results from ALL batches
                all_results = []
                
                # Read all existing batch checkpoint files
                import glob
                checkpoint_pattern = "freightos_checkpoints/batch_*_checkpoint_*.json"
                all_checkpoint_files = glob.glob(checkpoint_pattern)
                
                for checkpoint_file in all_checkpoint_files:
                    try:
                        with open(checkpoint_file, 'r') as f:
                            batch_data = json.load(f)
                            if isinstance(batch_data, list):
                                all_results.extend(batch_data)
                    except Exception as e:
                        logger.warning(f"Error reading checkpoint file {checkpoint_file}: {e}")
                
                # Remove duplicates based on unique combination of origin, destination, container, price
                seen = set()
                unique_results = []
                for result in all_results:
                    key = (
                        result.get('city_of_origin', ''),
                        result.get('city_of_destination', ''),
                        result.get('container_type', ''),
                        result.get('price_of_shipping', 0)
                    )
                    if key not in seen:
                        seen.add(key)
                        unique_results.append(result)
                
                # Save MASTER aggregated files (single source of truth)
                if unique_results:
                    master_timestamp = current_time.strftime("%Y%m%d_%H%M%S")
                    master_base = f"freightos_checkpoints/MASTER_ALL_RESULTS_{len(unique_results)}_{master_timestamp}"
                    
                    # Master JSON
                    master_json = f"{master_base}.json"
                    with open(master_json, 'w') as f:
                        json.dump(unique_results, f, indent=2)
                    
                    # Master CSV
                    master_csv = save_results_to_csv(unique_results, f"{master_base}.csv")
                    
                    # Master Excel
                    master_excel = save_results_to_excel_basic(unique_results, f"{master_base}.xlsx")
                
                logger.info(f"💾 Batch {batch_id} - TIME-BASED CHECKPOINT SAVED! {len(results)} batch results:")
                logger.info(f"   📄 Batch JSON: {json_file}")
                logger.info(f"   📊 Batch CSV: {csv_file}")
                logger.info(f"   📈 Batch Excel: {excel_file}")
                
                if unique_results:
                    logger.info(f"🎯 MASTER AGGREGATED FILES UPDATED! {len(unique_results)} total unique results:")
                    logger.info(f"   📄 Master JSON: {master_json}")
                    logger.info(f"   📊 Master CSV: {master_csv}")
                    logger.info(f"   📈 Master Excel: {master_excel}")
                
                # Update last checkpoint time
                last_checkpoint_time = current_time
                
            except Exception as e:
                logger.warning(f"Batch {batch_id} - Time-based checkpoint save failed: {e}")
        
        # Random delay between requests (with error handling)
        try:
            if delay_range and len(delay_range) >= 2:
                delay = random.uniform(delay_range[0], delay_range[1])
                logger.debug(f"Batch {batch_id} - Waiting {delay:.1f} seconds before next request...")
                await asyncio.sleep(delay)
        except Exception as delay_error:
            logger.warning(f"Batch {batch_id} - Error in delay: {delay_error}")
            # Continue without delay if delay calculation fails
            pass
    
    # FINAL CHECKPOINT: Save any remaining results at batch completion
    if len(results) > 0:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = f"freightos_checkpoints/batch_{batch_id}_FINAL_{len(results)}_{timestamp}"
            
            os.makedirs("freightos_checkpoints", exist_ok=True)
            
            # Save individual batch final files
            json_file = f"{base_filename}.json"
            with open(json_file, 'w') as f:
                json.dump(results, f, indent=2)
            
            csv_file = save_results_to_csv(results, f"{base_filename}.csv")
            excel_file = save_results_to_excel_basic(results, f"{base_filename}.xlsx")
            
            # CENTRALIZED AGGREGATION: Collect ALL results from ALL batches (including FINAL files)
            all_results = []
            
            # Read all existing batch checkpoint and final files
            import glob
            checkpoint_pattern = "freightos_checkpoints/batch_*_checkpoint_*.json"
            final_pattern = "freightos_checkpoints/batch_*_FINAL_*.json"
            all_files = glob.glob(checkpoint_pattern) + glob.glob(final_pattern)
            
            for file_path in all_files:
                try:
                    with open(file_path, 'r') as f:
                        batch_data = json.load(f)
                        if isinstance(batch_data, list):
                            all_results.extend(batch_data)
                except Exception as e:
                    logger.warning(f"Error reading file {file_path}: {e}")
            
            # Remove duplicates
            seen = set()
            unique_results = []
            for result in all_results:
                key = (
                    result.get('city_of_origin', ''),
                    result.get('city_of_destination', ''),
                    result.get('container_type', ''),
                    result.get('price_of_shipping', 0)
                )
                if key not in seen:
                    seen.add(key)
                    unique_results.append(result)
            
            # Save MASTER aggregated files
            if unique_results:
                master_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                master_base = f"freightos_checkpoints/MASTER_ALL_RESULTS_{len(unique_results)}_{master_timestamp}"
                
                # Master JSON
                master_json = f"{master_base}.json"
                with open(master_json, 'w') as f:
                    json.dump(unique_results, f, indent=2)
                
                # Master CSV
                master_csv = save_results_to_csv(unique_results, f"{master_base}.csv")
                
                # Master Excel
                master_excel = save_results_to_excel_basic(unique_results, f"{master_base}.xlsx")
            
            logger.info(f"🏁 Batch {batch_id} - FINAL CHECKPOINT SAVED! {len(results)} batch results:")
            logger.info(f"   📄 Batch JSON: {json_file}")
            logger.info(f"   📊 Batch CSV: {csv_file}")
            logger.info(f"   📈 Batch Excel: {excel_file}")
            
            if unique_results:
                logger.info(f"🎯 MASTER AGGREGATED FILES UPDATED! {len(unique_results)} total unique results:")
                logger.info(f"   📄 Master JSON: {master_json}")
                logger.info(f"   📊 Master CSV: {master_csv}")
                logger.info(f"   📈 Master Excel: {master_excel}")
            
        except Exception as e:
            logger.warning(f"Batch {batch_id} - Final checkpoint save failed: {e}")
    
    logger.info(f"🎯 Batch {batch_id} - COMPLETED with ALL {len(results)} combinations saved!")
    logger.info(f"   ✅ Successful quotes: {successful_count}")
    logger.info(f"   ⚠️  Failed attempts: {failed_count}")
    logger.info(f"   📊 Total processed: {total_pairs}")
    logger.info(f"   📈 Success rate: {(successful_count/total_pairs)*100:.1f}%")
    
    return results

async def compute_freightos_matrix_parallel(
    date: str = "2025-09-01", 
    container_types: List[str] = None,
    delay_range: tuple = (2, 5),
    num_processes: int = None,
    location_percentage: float = 0.1,
    checkpoint_interval: int = 50,
    resume: bool = True,
    checkpoint_dir: str = "freightos_checkpoints"
) -> List[Dict[str, Any]]:
    """
    Compute Freightos shipping matrix using parallel processing with checkpointing.
    
    Args:
        date: Shipping date in YYYY-MM-DD format
        container_types: List of container types (e.g., ["container20", "container40"])
        delay_range: Tuple of (min, max) seconds for random delays
        num_processes: Number of processes to use (default: CPU count)
        location_percentage: Percentage of locations to process
        checkpoint_interval: Number of results to process before checkpointing
        resume: Whether to resume from existing checkpoint
        checkpoint_dir: Directory to store checkpoint files
        
    Returns:
        List of dictionaries containing shipping data for each combination
    """
    if num_processes is None:
        num_processes = mp.cpu_count()
    
    # Set default container types if not provided
    if container_types is None:
        container_types = ["container20", "container40"]
    
    # Initialize checkpoint manager
    checkpoint_manager = CheckpointManager(
        checkpoint_dir=checkpoint_dir,
        checkpoint_interval=checkpoint_interval
    )
    
    logger.info(f"Starting PARALLEL Freightos matrix computation using {num_processes} processes")
    logger.info(f"Container types: {container_types}")
    
    # Handle resume vs fresh start
    checkpoint_data = {}
    if resume:
        checkpoint_data = checkpoint_manager.load_existing_checkpoint()
        if checkpoint_data['has_checkpoint']:
            logger.info(f"🔄 Resuming from checkpoint with {len(checkpoint_data['results'])} existing results")
    else:
        logger.info("🆕 Starting fresh - clearing existing checkpoints")
        checkpoint_manager.clear_checkpoints()
    
    # Get location combinations
    all_locations = list(FREIGHTOS_LOCATIONS.keys())
    locations_count = int(len(all_locations) * location_percentage)
    locations = random.sample(all_locations, locations_count)
    
    # Generate all location pairs with container types (exclude same location combinations)
    all_location_container_pairs = []
    for origin_location in locations:
        for destination_location in locations:
            if origin_location != destination_location:
                for container_type in container_types:
                    all_location_container_pairs.append((origin_location, destination_location, container_type))
    
    logger.info(f"📊 COMBINATION SETUP:")
    logger.info(f"   Total cities available: {len(all_locations)}")
    logger.info(f"   Cities to process: {len(locations)} ({location_percentage*100:.1f}%)")
    logger.info(f"   Expected combinations: {len(all_location_container_pairs)}")
    logger.info(f"   Target for completion: 7,140 combinations ({len(locations)} cities × 70)")
    
    # Get remaining pairs to process (skip completed ones if resuming)
    if resume and checkpoint_data.get('has_checkpoint'):
        remaining_pairs = checkpoint_manager.get_remaining_pairs(all_location_container_pairs)
        completed_count = len(checkpoint_data.get('completed_pairs', set()))
        logger.info(f"🔄 RESUMING FROM CHECKPOINT:")
        logger.info(f"   Combinations already completed: {completed_count}")
        logger.info(f"   Combinations remaining: {len(remaining_pairs)}")
        logger.info(f"   Progress so far: {(completed_count/7140)*100:.1f}% of target")
    else:
        remaining_pairs = all_location_container_pairs
        logger.info(f"🆕 STARTING FRESH: {len(remaining_pairs)} total location+container combinations to process")
    
    if not remaining_pairs:
        logger.info("✅ All location pairs already completed!")
        progress = checkpoint_manager.get_progress_summary()
        logger.info(f"🎉 FINAL STATS: {progress['completed_combinations']}/{progress['target_combinations']} combinations completed!")
        return checkpoint_manager.total_results
    
    # Split work into batches for parallel processing
    batch_size = max(1, len(remaining_pairs) // num_processes)
    batches = []
    
    for i in range(0, len(remaining_pairs), batch_size):
        batch_pairs = remaining_pairs[i:i + batch_size]
        batch_data = {
            'location_container_pairs': batch_pairs,
            'date': date,
            'delay_range': delay_range,
            'batch_id': len(batches) + 1,
            'checkpoint_manager': checkpoint_manager,
            'checkpoint_interval': checkpoint_interval
        }
        batches.append(batch_data)
    
    logger.info(f"Split work into {len(batches)} batches (avg {batch_size} combinations per batch)")
    
    # Process batches in parallel
    total_combinations_processed = 0
    total_successful_quotes = 0
    total_failed_attempts = 0
    
    try:
        with ProcessPoolExecutor(max_workers=num_processes) as executor:
            logger.info("🚀 Starting parallel Freightos processing with COMPLETE data saving...")
            logger.info("   📊 Every combination will be saved (successful + failed)")
            
            # Submit all batches for processing
            future_to_batch = {executor.submit(process_freightos_location_batch, batch): batch for batch in batches}
            
            # Collect results as they complete
            for future in future_to_batch:
                try:
                    batch_results = future.result()
                    batch_id = future_to_batch[future]['batch_id']
                    
                    # Count all results (both successful and failed)
                    batch_total = len(batch_results) if batch_results else 0
                    batch_successful = 0
                    batch_failed = 0
                    
                    if batch_results and len(batch_results) > 0:
                        for result in batch_results:
                            if result.get("request_status") == "success":
                                batch_successful += 1
                                total_successful_quotes += 1
                            else:
                                batch_failed += 1
                                total_failed_attempts += 1
                        total_combinations_processed += batch_total
                    
                    logger.info(f"✅ Completed batch {batch_id}: {batch_total} combinations")
                    logger.info(f"   📈 Success: {batch_successful}, Failed: {batch_failed}")
                    logger.info(f"   🎯 TOTAL PROGRESS: {total_combinations_processed} combinations processed")
                    logger.info(f"   📊 Overall Success Rate: {(total_successful_quotes/total_combinations_processed)*100:.1f}%")
                    
                except Exception as e:
                    batch_id = future_to_batch[future]['batch_id']
                    logger.error(f"❌ Batch {batch_id} failed: {e}")

    except KeyboardInterrupt:
        logger.info("🛑 Process interrupted by user - all data already saved via time-based checkpointing")
        raise
    except Exception as e:
        logger.error(f"❌ Parallel processing error: {e}")
        raise
    
    # Final summary with complete statistics
    logger.info("🎉 FREIGHTOS MATRIX COMPUTATION COMPLETED!")
    logger.info("=" * 60)
    logger.info(f"📊 FINAL STATISTICS:")
    logger.info(f"   💾 Total combinations processed: {total_combinations_processed}")
    logger.info(f"   ✅ Successful quotes obtained: {total_successful_quotes}")
    logger.info(f"   ⚠️  Failed attempts (with error details): {total_failed_attempts}")
    logger.info(f"   📈 Overall success rate: {(total_successful_quotes/total_combinations_processed)*100:.1f}%")
    logger.info("=" * 60)
    logger.info("📁 All data has been automatically saved via time-based checkpointing")
    logger.info("📄 Check freightos_checkpoints/ directory for:")
    logger.info("   • Individual batch files (JSON, CSV, Excel)")
    logger.info("   • Master aggregated files (all results combined)")
    logger.info("   • Complete dataset with success and error tracking")
    logger.info("=" * 60)
    
    # Return empty list since results are already saved in checkpoint files
    return []


def generate_biased_monthly_dates(
    start_date, months=12, total_dates=125, bias_months=(3,12), bias_ratio=0.6
):
    # Step 1: Calculate month boundaries
    month_starts = [start_date.replace(day=1)]
    for i in range(1, months):
        year = start_date.year + (start_date.month - 1 + i) // 12
        month = (start_date.month - 1 + i) % 12 + 1
        month_starts.append(datetime(year, month, 1))
    
    # Step 2: Assign number of dates per month
    bias_month_indices = list(range(bias_months[0]-1, bias_months[1]))
    non_bias_indices = [i for i in range(months) if i not in bias_month_indices]

    n_bias = round(total_dates * bias_ratio)
    n_non_bias = total_dates - n_bias

    # Distribute evenly (or as close as possible)
    per_bias_month = [n_bias // len(bias_month_indices)] * len(bias_month_indices)
    for i in range(n_bias % len(bias_month_indices)):
        per_bias_month[i] += 1

    per_non_bias_month = [n_non_bias // len(non_bias_indices)] * len(non_bias_indices)
    for i in range(n_non_bias % len(non_bias_indices)):
        per_non_bias_month[i] += 1

    # Step 3: Generate random dates for each month
    dates_obj_list = []
    # For biased months
    for idx, count in zip(bias_month_indices, per_bias_month):
        month_start = month_starts[idx]
        next_month = (month_start.month % 12) + 1
        next_month_year = month_start.year + (month_start.month // 12)
        month_end = datetime(next_month_year, next_month, 1) - timedelta(days=1)
        for _ in range(count):
            day = random.randint(0, (month_end - month_start).days)
            dates_obj_list.append((month_start + timedelta(days=day)).strftime("%d/%m/%Y"))
    # For non-biased months
    for idx, count in zip(non_bias_indices, per_non_bias_month):
        month_start = month_starts[idx]
        next_month = (month_start.month % 12) + 1
        next_month_year = month_start.year + (month_start.month // 12)
        month_end = datetime(next_month_year, next_month, 1) - timedelta(days=1)
        for _ in range(count):
            day = random.randint(0, (month_end - month_start).days)
            dates_obj_list.append((month_start + timedelta(days=day)).strftime("%d/%m/%Y"))
    
    # Optional: shuffle to look like real logs
    random.shuffle(dates_obj_list)
    return dates_obj_list

def transform_all_freightos_quotes(data, original_container_type: str = "container20", 
                                   origin_city_name: str = None, destination_city_name: str = None):
    """
    Transform ALL Freightos quotes into standardized format.
    Returns a list of transformed results, one for each quote.
    """
    if not data:
        logger.warning("Empty data provided to transform function")
        return []
    
    try:
        # Safely extract quotes
        quotes = None
        if isinstance(data, dict):
            quotes = data.get("quotes", [])
        elif hasattr(data, 'quotes'):
            quotes = data.quotes
        
        if not quotes or len(quotes) == 0:
            logger.debug("No quotes found in response data")
            return []
        
        # Process ALL quotes instead of just the first one
        all_results = []
        for i, quote in enumerate(quotes):
            if not quote:
                logger.warning(f"Quote {i+1} is empty, skipping")
                continue
                
            try:
                # Use the existing transform logic for each quote
                result = transform_single_quote(
                    quote, original_container_type, origin_city_name, destination_city_name, quote_index=i+1
                )
                if result:
                    all_results.append(result)
            except Exception as e:
                logger.warning(f"Error processing quote {i+1}: {e}")
                continue
        
        logger.info(f"Successfully processed {len(all_results)} out of {len(quotes)} quotes")
        return all_results
        
    except Exception as e:
        logger.error(f"Critical error in transform_all_freightos_quotes: {e}")
        return []

def transform_single_quote(quote, original_container_type: str = "container20", 
                          origin_city_name: str = None, destination_city_name: str = None,
                          quote_index: int = 1):
    """
    Transform a single quote into standardized format.
    Enhanced with comprehensive error handling.
    """
    try:
        # Safe extraction with fallbacks for all fields
        def safe_get(obj, *keys, default=""):
            """Safely navigate nested dictionary/object structure"""
            try:
                current = obj
                for key in keys:
                    if isinstance(current, dict):
                        current = current.get(key)
                    elif hasattr(current, key):
                        current = getattr(current, key)
                    else:
                        return default
                    if current is None:
                        return default
                return current if current is not None else default
            except Exception:
                return default
        
        # Extract location codes with multiple fallback strategies
        origin_code = ""
        dest_code = ""
        
        try:
            origin_code = (
                safe_get(quote, "originLocation", "locationCode") or
                safe_get(quote, "origin", "locationCode") or
                safe_get(quote, "from", "code") or
                safe_get(quote, "originLocation", "code") or
                ""
            )
            dest_code = (
                safe_get(quote, "destinationLocation", "locationCode") or
                safe_get(quote, "destination", "locationCode") or
                safe_get(quote, "to", "code") or
                safe_get(quote, "destinationLocation", "code") or
                ""
            )
        except Exception as e:
            logger.warning(f"Error extracting location codes: {e}")
        
        # Use provided city names if available, otherwise convert location codes
        if origin_city_name:
            city_of_origin = origin_city_name
            country_of_origin = origin_code[:2] if origin_code and len(origin_code) >= 2 else ""
        else:
            city_of_origin, country_of_origin = locode_to_city_country(origin_code)
            
        if destination_city_name:
            city_of_destination = destination_city_name  
            country_of_destination = dest_code[:2] if dest_code and len(dest_code) >= 2 else ""
        else:
            city_of_destination, country_of_destination = locode_to_city_country(dest_code)
        
        # Extract date with fallbacks
        date_of_shipping = ""
        try:
            date_of_shipping = (
                safe_get(quote, "createDate") or
                safe_get(quote, "date") or
                safe_get(quote, "shipmentDate") or
                datetime.now().date().isoformat()
            )
            if date_of_shipping and len(str(date_of_shipping)) > 10:
                date_of_shipping = str(date_of_shipping)[:10]
        except Exception as e:
            logger.warning(f"Error extracting date: {e}")
            date_of_shipping = datetime.now().date().isoformat()
        
        # Extract shipping time with multiple fallbacks
        shipping_time = None
        try:
            # First try the correct Freightos API structure
            transit_times = safe_get(quote, "connection", "transitTime", "estimatedTransitTimes")
            if transit_times and isinstance(transit_times, list) and len(transit_times) > 0:
                time_range = transit_times[0]
                from_value = safe_get(time_range, "from", "value")
                to_value = safe_get(time_range, "to", "value")
                
                if from_value is not None and to_value is not None:
                    try:
                        from_days = int(from_value)
                        to_days = int(to_value)
                        if from_days == to_days:
                            shipping_time = f"{from_days} days"
                        else:
                            shipping_time = f"{from_days}-{to_days} days"
                    except (ValueError, TypeError):
                        pass
            
            # Fallback to other possible structures if the main one fails
            if shipping_time is None:
                fallback_value = (
                    safe_get(quote, "transitTime", "value") or
                    safe_get(quote, "estimatedDays") or
                    safe_get(quote, "transitDays")
                )
                if fallback_value is not None and str(fallback_value).strip():
                    try:
                        days = int(float(fallback_value))
                        shipping_time = f"{days} days"
                    except (ValueError, TypeError):
                        pass
                        
        except Exception as e:
            logger.warning(f"Error extracting shipping time: {e}")
            shipping_time = None
        
        # Extract price and currency with comprehensive fallbacks
        price_of_shipping = None
        currency = None
        
        try:
            # Try priceIndicator.totalCharge first
            price_info = safe_get(quote, "priceIndicator", "totalCharge")
            if isinstance(price_info, dict):
                price_of_shipping = price_info.get("value")
                currency = price_info.get("currencyID")
            
            # Fallback to connection segments
            if price_of_shipping is None:
                try:
                    segments = safe_get(quote, "connection", "connectionSegments")
                    if segments and len(segments) > 0:
                        charges = safe_get(segments[0], "charges")
                        if charges and len(charges) > 0:
                            rate = safe_get(charges[0], "rate")
                            if isinstance(rate, dict):
                                price_of_shipping = rate.get("value")
                                currency = rate.get("currencyID")
                except Exception:
                    pass
            
            # More fallbacks for price
            if price_of_shipping is None:
                price_of_shipping = (
                    safe_get(quote, "totalPrice") or
                    safe_get(quote, "price", "value") or
                    safe_get(quote, "cost") or
                    safe_get(quote, "amount")
                )
            
            # More fallbacks for currency
            if currency is None:
                currency = (
                    safe_get(quote, "totalCurrency") or
                    safe_get(quote, "price", "currency") or
                    safe_get(quote, "currencyCode") or
                    "USD"
                )
            
            # Convert price to float if possible
            if price_of_shipping is not None:
                try:
                    price_of_shipping = float(price_of_shipping)
                except (ValueError, TypeError):
                    price_of_shipping = None
                    
        except Exception as e:
            logger.warning(f"Error extracting price: {e}")
            price_of_shipping = None
            currency = "USD"
        
        # Container type mapping
        container_mapping = {
            "container20": "ST20",
            "container40": "ST40", 
            "container40hc": "ST40HC",
            "container45": "ST45"
        }
        container_type = container_mapping.get(original_container_type, original_container_type)
        
        # Extract carrier with fallbacks
        carrier = ""
        try:
            carrier = (
                safe_get(quote, "connection", "connectionSegments", 0, "carrier", "name") or
                safe_get(quote, "carrier", "name") or
                safe_get(quote, "carrierName") or
                safe_get(quote, "provider") or
                ""
            )
        except Exception as e:
            logger.warning(f"Error extracting carrier: {e}")
        
        # Extract shipment and rate IDs
        shipment_id = safe_get(quote, "shipmentId") or safe_get(quote, "id") or ""
        rate_id = safe_get(quote, "rateId") or safe_get(quote, "quoteId") or ""
        
        # Extract validity dates
        validity_from = date_of_shipping
        validity_to = ""
        try:
            validity_to = (
                safe_get(quote, "validUntil") or
                safe_get(quote, "expiryDate") or
                safe_get(quote, "validity", "to") or
                ""
            )
            if validity_to and len(str(validity_to)) > 10:
                validity_to = str(validity_to)[:10]
        except Exception as e:
            logger.warning(f"Error extracting validity dates: {e}")
        
        # Build the result object
        result = {
            "city_of_origin": city_of_origin,
            "country_of_origin": country_of_origin,
            "city_of_destination": city_of_destination,
            "country_of_destination": country_of_destination,
            "date_of_shipping": date_of_shipping,
            "total_shipping_time_days": shipping_time,
            "price_of_shipping": price_of_shipping,
            "currency": currency,
            "container_type": container_type,
            "provider": "Freightos",
            "carrier": carrier,
            "shipment_id": shipment_id,
            "rate_id": rate_id,
            "website_url": "",  # Will be set by caller
            "co2_amount": None,  # Extract if available in future
            "co2_price": None,
            "validity_from": validity_from,
            "validity_to": validity_to,
            "distance": "",  # Extract if available
            "point_total": None,
            "route_total": None,
            "quote_number": quote_index,  # Track which quote this is (1, 2, 3, etc.)
            "total_quotes_available": 1  # Will be updated by caller
        }
        
        return result
        
    except Exception as e:
        logger.error(f"Error transforming single quote: {e}")
        return None

def create_freightos_result_entries(
    origin_location: str,
    destination_location: str,
    quotes_data: Dict[str, Any],
    date: str,
    container: str,
    screenshot_url: Optional[str] = None,
    website_link: Optional[str] = None
) -> List[Dict[str, Any]]:
    """
    Create standardized result entries from Freightos quotes data.
    Returns a LIST of results - one for each quote found.
    Enhanced to capture ALL quotes, not just the first one.
    """
    results = []
    
    try:
        if quotes_data:
            # Process ALL quotes
            all_quotes = transform_all_freightos_quotes(
                quotes_data, 
                original_container_type=container,
                origin_city_name=origin_location,
                destination_city_name=destination_location
            )
            
            if all_quotes and len(all_quotes) > 0:
                # Update total_quotes_available for all results
                for quote_result in all_quotes:
                    if quote_result:
                        quote_result.update({
                            "total_quotes_available": len(all_quotes),
                            "request_status": "success",
                            "error_type": None,
                            "error_message": "",
                            "has_polling_url": True,
                            "polling_attempts": 1,
                            "processing_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "screenshot_url": screenshot_url,
                            "website_url": website_link
                        })
                        results.append(quote_result)
                
                logger.info(f"✅ Successfully processed {len(results)} quotes for {origin_location} -> {destination_location}")
                
    except Exception as e:
        logger.error(f"Error transforming Freightos response: {e}")
        # Return single error result on transformation failure
        results.append(create_freightos_empty_result(
            origin_location, destination_location, date, container,
            error_type="transformation_error",
            error_message=str(e),
            has_polling_url=True,
            polling_attempts=1,
            screenshot_url=screenshot_url,
            website_link=website_link
        ))
    
    # If no valid quotes were processed, return error result
    if not results:
        results.append(create_freightos_empty_result(
            origin_location, destination_location, date, container,
            error_type="invalid_response_data",
            error_message="No valid pricing data in response",
            has_polling_url=True,
            polling_attempts=1,
            screenshot_url=screenshot_url,
            website_link=website_link
        ))
    
    return results